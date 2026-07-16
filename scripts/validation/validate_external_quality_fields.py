#!/usr/bin/env python3
"""Validate a contiguous field range in a project operating-quality report.

All customer workbooks are supplied through command-line arguments. The script
does not embed local workbook names or customer/entity exceptions, so it can be
kept as a reusable governance artifact while all sensitive data stays local.
"""

from __future__ import annotations

import argparse
import json
import math
import re
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable
from xml.etree import ElementTree as ET
from zipfile import ZipFile

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter


AMOUNT_TOLERANCE = 0.01
RATIO_TOLERANCE = 0.000001
STATUS_PASS = "通过"
STATUS_FAIL = "失败"
STATUS_BLOCKED = "缺源阻塞"
STATUS_REVIEW = "待口径确认"
EXCEL_ERROR_VALUES = {
    "#DIV/0!",
    "#N/A",
    "#NAME?",
    "#NULL!",
    "#NUM!",
    "#REF!",
    "#VALUE!",
}

XML_NS = {
    "m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main",
    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships",
}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--template", type=Path, required=True)
    parser.add_argument("--indicator-list", type=Path, required=True)
    parser.add_argument("--project-master", type=Path, required=True)
    parser.add_argument("--saturation-ledger", type=Path, required=True)
    parser.add_argument("--projection-ledger", type=Path, required=True)
    parser.add_argument("--business-aging", type=Path, required=True)
    parser.add_argument("--receivable-aging", type=Path, required=True)
    parser.add_argument("--half-cash-source", type=Path, required=True)
    parser.add_argument("--accrual-source", type=Path)
    parser.add_argument("--current-accrual-source", type=Path)
    parser.add_argument("--current-half-cash-source", type=Path)
    parser.add_argument("--current-business-aging", type=Path)
    parser.add_argument("--current-receivable-aging", type=Path)
    parser.add_argument("--occupancy-source", type=Path)
    parser.add_argument("--amortization-source", type=Path)
    parser.add_argument("--cumulative-accrual-source", type=Path)
    parser.add_argument("--cumulative-half-cash-source", type=Path)
    parser.add_argument(
        "--cumulative-financial-policy",
        choices=("direct-range", "sum-period-sources"),
        default="direct-range",
        help="Use direct entry-to-report sources or sum first/current period sources.",
    )
    parser.add_argument(
        "--report-period",
        help="Report period in YYYYMM form; required for the full field set.",
    )
    parser.add_argument(
        "--confirmed-absent-project-code",
        action="append",
        default=[],
        help="Project code explicitly confirmed absent from the report-period projection ledger.",
    )
    parser.add_argument(
        "--field-set",
        choices=(
            "original",
            "external-full",
            "remaining-report-formulas",
            "all-report-formulas",
        ),
        default="original",
        help="Select the original, extended, remaining, or all report-formula fields.",
    )
    parser.add_argument(
        "--cash-comparative-report",
        type=Path,
        help="Optional project report carrying current cumulative cash and MoM.",
    )
    parser.add_argument(
        "--missing-row-policy",
        choices=("blocked", "zero"),
        default="blocked",
        help="How to treat a project that is absent from an otherwise present source.",
    )
    parser.add_argument(
        "--cumulative-policy",
        choices=("indicator-year", "entry-month", "both"),
        default="both",
        help="Validate cumulative cash from January, from entry month, or report both.",
    )
    parser.add_argument("--output", type=Path, required=True)
    parser.add_argument("--summary-json", type=Path)
    return parser.parse_args()


def norm_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number(value: Any) -> float:
    if value is None or value == "":
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def comparable_number(value: Any) -> float | None:
    if value is None or value == "":
        return None
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def near_zero(value: float) -> bool:
    return abs(value) < AMOUNT_TOLERANCE


def safe_ratio(numerator: float, denominator: float) -> float:
    if near_zero(denominator):
        return 0.0
    return numerator / denominator


def compare(actual: Any, expected: Any, kind: str) -> tuple[bool, float | None]:
    if kind == "text":
        return norm_text(actual) == norm_text(expected), None
    if kind == "year":
        actual_text = norm_text(actual)
        expected_text = norm_text(expected)
        return actual_text == expected_text, None
    actual_num = comparable_number(actual)
    expected_num = comparable_number(expected)
    if actual_num is None or expected_num is None:
        return actual_num is None and expected_num is None, None
    diff = actual_num - expected_num
    tolerance = RATIO_TOLERANCE if kind == "ratio" else AMOUNT_TOLERANCE
    return abs(diff) <= tolerance, diff


def read_sheet(path: Path, data_only: bool = True):
    workbook = load_workbook(path, read_only=False, data_only=data_only)
    return workbook.active


def rows_by_key(
    path: Path, start_row: int, key_index: int, data_only: bool = True
) -> tuple[Any, dict[str, list[tuple[Any, ...]]]]:
    sheet = read_sheet(path, data_only=data_only)
    result: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    for row in sheet.iter_rows(min_row=start_row, values_only=True):
        key = norm_text(row[key_index] if key_index < len(row) else None)
        if key:
            result[key].append(row)
    return sheet, result


def add_months(year: int, month: int, offset: int) -> tuple[int, int]:
    serial = year * 12 + (month - 1) + offset
    return serial // 12, serial % 12 + 1


def parse_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    text = norm_text(value)
    for fmt in ("%Y-%m-%d", "%Y/%m/%d", "%Y-%m", "%Y/%m"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def normalized_entry_month(value: Any) -> str:
    parsed = parse_date(value)
    if parsed is None:
        return ""
    offset = 1 if parsed.day >= 15 else 0
    year, month = add_months(parsed.year, parsed.month, offset)
    return f"{year:04d}-{month:02d}"


def shifted_period(value: Any, months: int) -> str:
    period = normalized_entry_month(value)
    if not period:
        return ""
    year, month = (int(part) for part in period.split("-"))
    year, month = add_months(year, month, months)
    return f"{year:04d}-{month:02d}"


def normalized_project_code(value: Any) -> str:
    """Normalize reports that inconsistently omit the leading project type."""
    text = norm_text(value)
    if len(text) > 1 and text[0] in {"P", "D", "Z"} and text[1].isdigit():
        return text[1:]
    return text


def year_from_period(value: Any) -> str:
    text = norm_text(value)
    match = re.match(r"^(\d{4})", text)
    return match.group(1) if match else ""


def normalized_period(value: Any) -> str:
    digits = re.sub(r"\D", "", norm_text(value))
    return digits[:6] if len(digits) >= 6 else ""


def period_reached(value: Any, report_period: str) -> bool:
    period = normalized_period(value)
    return bool(period and period <= report_period)


def xlsx_cells_without_styles(path: Path) -> dict[str, dict[str, str]]:
    """Read cell values directly from OOXML, bypassing malformed style XML."""
    with ZipFile(path) as archive:
        shared: list[str] = []
        if "xl/sharedStrings.xml" in archive.namelist():
            root = ET.fromstring(archive.read("xl/sharedStrings.xml"))
            shared = [
                "".join(node.text or "" for node in item.iterfind(".//m:t", XML_NS))
                for item in root.findall("m:si", XML_NS)
            ]
        workbook = ET.fromstring(archive.read("xl/workbook.xml"))
        relationships = ET.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
        relation_map = {node.attrib["Id"]: node.attrib["Target"] for node in relationships}
        result: dict[str, dict[str, str]] = {}
        sheets = workbook.find("m:sheets", XML_NS)
        if sheets is None:
            return result
        for sheet in sheets:
            name = sheet.attrib["name"]
            relation_id = sheet.attrib[f"{{{XML_NS['r']}}}id"]
            target = relation_map[relation_id]
            if not target.startswith("xl/"):
                target = "xl/" + target.lstrip("/")
            sheet_root = ET.fromstring(archive.read(target))
            values: dict[str, str] = {}
            for cell in sheet_root.iterfind(".//m:c", XML_NS):
                reference = cell.attrib["r"]
                cell_type = cell.attrib.get("t")
                value_node = cell.find("m:v", XML_NS)
                value = ""
                if cell_type == "inlineStr":
                    value = "".join(
                        node.text or "" for node in cell.iterfind(".//m:t", XML_NS)
                    )
                elif value_node is not None:
                    raw = value_node.text or ""
                    value = shared[int(raw)] if cell_type == "s" else raw
                if value != "":
                    values[reference] = value
            result[name] = values
        return result


def row_number(reference: str) -> int:
    match = re.search(r"\d+", reference)
    return int(match.group()) if match else 0


def indicator_evidence(path: Path) -> dict[str, bool]:
    workbook = xlsx_cells_without_styles(path)
    all_values = [value for sheet in workbook.values() for value in sheet.values()]
    required = {
        "累计回收现金流": False,
        "累计回收现金流_项目": False,
        "年度饱和收入_打折前": False,
    }
    for name in required:
        required[name] = any(value == name for value in all_values)
    return required


def template_evidence(path: Path) -> dict[str, Any]:
    workbook = load_workbook(path, read_only=False, data_only=True)
    for sheet in workbook.worksheets:
        values = {
            norm_text(cell.value): cell.coordinate
            for row in sheet.iter_rows(min_row=1, max_row=min(sheet.max_row, 5))
            for cell in row
            if cell.value is not None
        }
        if "合资公司名称" in values and "1个月时间" in values:
            return {
                "sheet": sheet.title,
                "start": values["合资公司名称"],
                "one_month_time": values["1个月时间"],
                "has_recovery_rate": any("回款率" in key for key in values),
            }
    return {"sheet": None, "start": None, "one_month_time": None, "has_recovery_rate": False}


def first_or_none(items: list[tuple[Any, ...]]) -> tuple[Any, ...] | None:
    return items[0] if len(items) == 1 else None


def sum_index(items: Iterable[tuple[Any, ...]], index: int) -> float:
    total = 0.0
    for item in items:
        if index < len(item):
            value = number(item[index])
            if not math.isnan(value):
                total += value
    return total


def find_header_row(sheet: Any, required: str, max_rows: int = 10) -> int:
    for row_index in range(1, min(max_rows, sheet.max_row) + 1):
        if any(norm_text(cell.value) == required for cell in sheet[row_index]):
            return row_index
    raise ValueError(f"Required header not found: {required}")


def build_cash_comparative(path: Path | None) -> dict[str, tuple[float, float, float]]:
    if path is None:
        return {}
    sheet = read_sheet(path, data_only=True)
    result: dict[str, tuple[float, float, float]] = {}
    for row in sheet.iter_rows(min_row=3, values_only=True):
        code = norm_text(row[1] if len(row) > 1 else None)
        if not code:
            continue
        cumulative = number(row[10]) * 10000
        mom = number(row[11])
        if math.isnan(cumulative) or math.isnan(mom):
            continue
        if near_zero(cumulative):
            monthly = 0.0
        elif abs(mom) <= RATIO_TOLERANCE:
            monthly = cumulative
        elif abs(1 + mom) <= RATIO_TOLERANCE:
            monthly = math.nan
        else:
            monthly = cumulative - cumulative / (1 + mom)
        result[code] = (cumulative, mom, monthly)
    return result


def output_value(value: Any) -> Any:
    if isinstance(value, float) and math.isnan(value):
        return None
    if isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_VALUES:
        return f"报表错误值:{value.strip().upper()}"
    return value


def is_excel_error(value: Any) -> bool:
    return isinstance(value, str) and value.strip().upper() in EXCEL_ERROR_VALUES


def main() -> int:
    args = parse_args()
    input_paths = [
        args.report,
        args.template,
        args.indicator_list,
        args.project_master,
        args.saturation_ledger,
        args.projection_ledger,
        args.business_aging,
        args.receivable_aging,
        args.half_cash_source,
    ]
    extended_paths = [
        args.accrual_source,
        args.current_business_aging,
        args.current_receivable_aging,
        args.occupancy_source,
        args.amortization_source,
    ]
    if args.field_set in {"external-full", "all-report-formulas"}:
        if not args.report_period or not re.fullmatch(r"\d{6}", args.report_period):
            raise ValueError(f"{args.field_set} requires --report-period YYYYMM")
    if args.field_set == "external-full":
        if any(path is None for path in extended_paths):
            raise ValueError(
                "external-full requires accrual, current, occupancy, and "
                "amortization source arguments"
            )
        input_paths.extend(path for path in extended_paths if path is not None)
        for optional_path in (
            args.current_accrual_source,
            args.current_half_cash_source,
            args.cumulative_accrual_source,
            args.cumulative_half_cash_source,
        ):
            if optional_path is not None:
                input_paths.append(optional_path)
    if args.cash_comparative_report:
        input_paths.append(args.cash_comparative_report)
    missing_files = [str(path) for path in input_paths if not path.exists()]
    if missing_files:
        raise FileNotFoundError("Missing inputs: " + ", ".join(missing_files))

    report_sheet = read_sheet(args.report, data_only=True)
    project_sheet, project_map = rows_by_key(args.project_master, 2, 0)
    project_name_map: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    for project_rows in project_map.values():
        for project_row in project_rows:
            project_name = norm_text(project_row[1] if len(project_row) > 1 else None)
            if project_name:
                project_name_map[project_name].append(project_row)
    saturation_sheet, saturation_map = rows_by_key(args.saturation_ledger, 2, 8)
    projection_sheet, projection_map = rows_by_key(args.projection_ledger, 4, 2)
    business_sheet, business_map = rows_by_key(args.business_aging, 5, 1)
    receivable_sheet, receivable_map = rows_by_key(args.receivable_aging, 3, 4)
    half_cash_sheet, half_cash_map = rows_by_key(args.half_cash_source, 6, 1)
    cash_comparative = build_cash_comparative(args.cash_comparative_report)

    accrual_map: dict[str, list[tuple[Any, ...]]] = {}
    current_accrual_map: dict[str, list[tuple[Any, ...]]] = {}
    current_half_cash_map: dict[str, list[tuple[Any, ...]]] = {}
    current_business_map: dict[str, list[tuple[Any, ...]]] = {}
    current_receivable_map: dict[str, list[tuple[Any, ...]]] = {}
    cumulative_accrual_map: dict[str, list[tuple[Any, ...]]] = {}
    cumulative_half_cash_map: dict[str, list[tuple[Any, ...]]] = {}
    occupancy_map: dict[str, list[tuple[Any, ...]]] = defaultdict(list)
    amortization_map: dict[str, list[tuple[Any, ...]]] = {}
    if args.field_set == "external-full":
        accrual_sheet, accrual_map = rows_by_key(args.accrual_source, 6, 1)
        if args.current_accrual_source:
            current_accrual_sheet, current_accrual_map = rows_by_key(
                args.current_accrual_source, 6, 1
            )
        if args.current_half_cash_source:
            current_half_cash_sheet, current_half_cash_map = rows_by_key(
                args.current_half_cash_source, 6, 1
            )
        current_business_sheet, current_business_map = rows_by_key(
            args.current_business_aging, 5, 1
        )
        current_receivable_sheet, current_receivable_map = rows_by_key(
            args.current_receivable_aging, 3, 4
        )
        occupancy_sheet = read_sheet(args.occupancy_source, data_only=True)
        for row in occupancy_sheet.iter_rows(min_row=3, values_only=True):
            code = normalized_project_code(row[3] if len(row) > 3 else None)
            if code:
                occupancy_map[code].append(row)
        amortization_sheet, amortization_map = rows_by_key(
            args.amortization_source, 3, 2
        )
        if args.cumulative_accrual_source:
            cumulative_accrual_sheet, cumulative_accrual_map = rows_by_key(
                args.cumulative_accrual_source, 6, 1
            )
        if args.cumulative_half_cash_source:
            cumulative_half_cash_sheet, cumulative_half_cash_map = rows_by_key(
                args.cumulative_half_cash_source, 6, 1
            )

    # Force structural checks so a wrong workbook fails loudly.
    find_header_row(project_sheet, "立项编码")
    find_header_row(saturation_sheet, "业绩认定年月")
    find_header_row(projection_sheet, "营业收入")
    find_header_row(business_sheet, "应收余额")
    find_header_row(receivable_sheet, "大业主应收金额 （单位：元）")
    find_header_row(half_cash_sheet, "累计回收现金流")
    if args.field_set == "external-full":
        find_header_row(accrual_sheet, "立项编码")
        if args.current_accrual_source:
            find_header_row(current_accrual_sheet, "立项编码")
        if args.current_half_cash_source:
            find_header_row(current_half_cash_sheet, "累计回收现金流")
        find_header_row(current_business_sheet, "应收余额")
        find_header_row(current_receivable_sheet, "大业主应收金额 （单位：元）")
        find_header_row(occupancy_sheet, "项目编号")
        find_header_row(amortization_sheet, "项目")
        if args.cumulative_accrual_source:
            find_header_row(cumulative_accrual_sheet, "立项编码")
        if args.cumulative_half_cash_source:
            find_header_row(cumulative_half_cash_sheet, "累计回收现金流")

    template_check = template_evidence(args.template)
    indicator_check = indicator_evidence(args.indicator_list)
    if not template_check["sheet"] or not all(indicator_check.values()):
        raise ValueError(
            f"Template/indicator evidence incomplete: {template_check}, {indicator_check}"
        )

    saturation_period_available = True
    projection_period_available = True
    cumulative_financial_available = False
    if args.field_set == "external-full":
        saturation_period_available = any(
            normalized_period(row[2]) == args.report_period
            for rows in saturation_map.values()
            for row in rows
        )
        projection_period_available = any(
            normalized_period(row[0]) == args.report_period
            for rows in projection_map.values()
            for row in rows
        )
        if bool(args.cumulative_accrual_source) != bool(
            args.cumulative_half_cash_source
        ):
            raise ValueError(
                "Provide both cumulative accrual and cumulative half-cash sources, or neither"
            )
        if args.cumulative_financial_policy == "sum-period-sources":
            if not args.current_accrual_source or not args.current_half_cash_source:
                raise ValueError(
                    "sum-period-sources requires current accrual and half-cash sources"
                )
            cumulative_financial_available = True
        else:
            cumulative_financial_available = bool(args.cumulative_accrual_source)

    field_specs = [
        (11, "K", "合资公司名称", "text", "项目主数据"),
        (12, "L", "股权比例", "ratio", "项目主数据"),
        (13, "M", "拓展年份", "year", "年饱和台账"),
        (14, "N", "年饱和", "amount", "年饱和台账"),
        (15, "O", "营业收入", "amount", "投模台账"),
        (16, "P", "半收付收入", "amount", "投模台账"),
        (17, "Q", "权责税前利润", "amount", "投模台账"),
        (18, "R", "半收付税前利润", "amount", "投模台账"),
        (19, "S", "营业成本（含增值税附加）", "amount", "投模台账"),
        (20, "T", "自有成本", "amount", "投模台账"),
        (21, "U", "外包成本", "amount", "投模台账"),
        (22, "V", "能耗成本", "amount", "投模台账"),
        (23, "W", "回款率", "ratio", "报表公式"),
        (24, "X", "投前测算入住率", "ratio", "投模台账"),
        (25, "Y", "增值税费附加", "amount", "投模台账"),
        (26, "Z", "管理费", "amount", "投模台账"),
        (27, "AA", "年终奖励/月", "amount", "投模台账"),
        (28, "AB", "带资投入费用", "amount", "投模台账"),
        (29, "AC", "市场营销费", "amount", "投模台账"),
        (30, "AD", "调整后营业收入", "amount", "报表公式"),
        (31, "AE", "调整后半收付收入", "amount", "报表公式"),
        (32, "AF", "调整后权责税前利润", "amount", "报表公式"),
        (33, "AG", "调整后半收付税前利润", "amount", "报表公式"),
        (34, "AH", "权责税前利润率", "ratio", "报表公式"),
        (35, "AI", "半收付税前利润率", "ratio", "报表公式"),
        (36, "AJ", "营业成本（不含带资）", "amount", "报表公式"),
        (37, "AK", "营业成本", "amount", "报表公式"),
        (38, "AL", "1个月时间", "text", "项目主数据"),
        (39, "AM", "累计回收现金流", "amount", "半收付底表"),
        (40, "AN", "应收账款余额", "amount", "账龄底表"),
        (41, "AO", "未到账期金额", "amount", "账龄底表"),
        (42, "AP", "回款率", "ratio", "报表公式"),
    ]
    if args.field_set == "external-full":
        field_specs = [item for item in field_specs if item[0] <= 37]
        field_specs.extend(
            [
                (63, "BK", "首年时间", "text", "项目主数据"),
                (64, "BL", "首年营业收入", "amount", "首年权责底表"),
                (65, "BM", "首年半收付收入", "amount", "首年半收付底表"),
                (66, "BN", "首年权责税前利润", "amount", "报表公式"),
                (67, "BO", "首年半收付税前利润", "amount", "报表公式"),
                (68, "BP", "首年营业成本", "amount", "首年权责底表"),
                (69, "BQ", "首年自有成本", "amount", "首年权责底表"),
                (70, "BR", "首年外包成本", "amount", "首年权责底表"),
                (71, "BS", "首年能耗成本", "amount", "首年权责底表"),
                (72, "BT", "首年累计回收现金流", "amount", "首年半收付底表"),
                (73, "BU", "首年应收余额", "amount", "首年账龄底表"),
                (74, "BV", "首年未到账期", "amount", "首年账龄底表"),
                (75, "BW", "首年回款率", "ratio", "报表公式"),
                (76, "BX", "首年带资投入费用", "amount", "摊销分析底表"),
                (77, "BY", "首年实际运营入住率", "ratio", "大额欠费分析"),
                (78, "BZ", "首年营业收入偏差", "amount", "报表公式"),
                (79, "CA", "首年半收付收入偏差", "amount", "报表公式"),
                (80, "CB", "首年权责税前利润偏差", "amount", "报表公式"),
                (81, "CC", "首年半收付税前利润偏差", "amount", "报表公式"),
                (82, "CD", "首年权责税前利润率偏差", "ratio", "报表公式"),
                (83, "CE", "首年半收付税前利润率偏差", "ratio", "报表公式"),
                (84, "CF", "首年营业成本偏差", "amount", "报表公式"),
                (85, "CG", "首年带资使用偏差", "amount", "报表公式"),
                (86, "CH", "首年自有成本偏差", "amount", "报表公式"),
                (87, "CI", "首年外包成本偏差", "amount", "报表公式"),
                (88, "CJ", "首年能耗成本偏差", "amount", "报表公式"),
                (89, "CK", "首年回款率偏差", "ratio", "报表公式"),
                (90, "CL", "进场后累计营业收入", "amount", "权责底表"),
                (91, "CM", "进场后累计半收付收入", "amount", "半收付底表"),
                (92, "CN", "进场后累计权责税前利润", "amount", "报表公式"),
                (93, "CO", "进场后累计半收付税前利润", "amount", "报表公式"),
                (94, "CP", "进场后累计营业成本", "amount", "权责底表"),
                (95, "CQ", "进场后累计自有成本", "amount", "权责底表"),
                (96, "CR", "进场后累计外包成本", "amount", "权责底表"),
                (97, "CS", "进场后累计能耗成本", "amount", "权责底表"),
                (98, "CT", "进场后累计回收现金流", "amount", "半收付底表"),
                (99, "CU", "进场后累计应收余额", "amount", "账龄底表"),
                (100, "CV", "进场后累计未到账期", "amount", "账龄底表"),
                (101, "CW", "进场后累计综合回款率", "ratio", "报表公式"),
            ]
        )
    elif args.field_set == "remaining-report-formulas":
        # Fields outside K:AK and BK:CW that are derived at report layer.
        # Row 5 of the template is only a hint: date shifts and the two
        # six-month profit fields are also formulas according to row 3.
        field_specs = [
            (9, "I", "进场时间", "text", "项目主数据+换月规则"),
            (38, "AL", "1个月时间", "text", "报表公式"),
            (42, "AP", "1个月回款率", "ratio", "报表公式"),
            (43, "AQ", "3个月时间", "text", "报表公式"),
            (47, "AU", "3个月回款率", "ratio", "报表公式"),
            (48, "AV", "6个月时间", "text", "报表公式"),
            (51, "AY", "6个月权责税前利润", "amount", "报表公式"),
            (52, "AZ", "6个月半收付税前利润", "amount", "报表公式"),
            (60, "BH", "6个月回款率", "ratio", "报表公式"),
        ]
    elif args.field_set == "all-report-formulas":
        field_specs = [
            (9, "I", "进场时间", "text", "项目主数据+换月规则"),
            (23, "W", "投模回款率", "ratio", "报表公式"),
            (30, "AD", "调整后营业收入", "amount", "报表公式"),
            (31, "AE", "调整后半收付收入", "amount", "报表公式"),
            (32, "AF", "调整后权责税前利润", "amount", "报表公式"),
            (33, "AG", "调整后半收付税前利润", "amount", "报表公式"),
            (34, "AH", "调整后权责税前利润率", "ratio", "报表公式"),
            (35, "AI", "调整后半收付税前利润率", "ratio", "报表公式"),
            (36, "AJ", "调整后营业成本不含带资", "amount", "报表公式"),
            (37, "AK", "调整后营业成本", "amount", "报表公式"),
            (38, "AL", "1个月时间", "text", "报表公式"),
            (42, "AP", "1个月回款率", "ratio", "报表公式"),
            (43, "AQ", "3个月时间", "text", "报表公式"),
            (47, "AU", "3个月回款率", "ratio", "报表公式"),
            (48, "AV", "6个月时间", "text", "报表公式"),
            (51, "AY", "6个月权责税前利润", "amount", "报表公式"),
            (52, "AZ", "6个月半收付税前利润", "amount", "报表公式"),
            (60, "BH", "6个月回款率", "ratio", "报表公式"),
            (63, "BK", "首年时间", "text", "报表公式"),
            (66, "BN", "首年权责税前利润", "amount", "报表公式"),
            (67, "BO", "首年半收付税前利润", "amount", "报表公式"),
            (75, "BW", "首年回款率", "ratio", "报表公式"),
            (78, "BZ", "首年营业收入偏差", "amount", "报表公式"),
            (79, "CA", "首年半收付收入偏差", "amount", "报表公式"),
            (80, "CB", "首年权责税前利润偏差", "amount", "报表公式"),
            (81, "CC", "首年半收付税前利润偏差", "amount", "报表公式"),
            (82, "CD", "首年权责税前利润率偏差", "ratio", "报表公式"),
            (83, "CE", "首年半收付税前利润率偏差", "ratio", "报表公式"),
            (84, "CF", "首年营业成本偏差", "amount", "报表公式"),
            (85, "CG", "首年带资使用偏差", "amount", "报表公式"),
            (86, "CH", "首年自有成本偏差", "amount", "报表公式"),
            (87, "CI", "首年外包成本偏差", "amount", "报表公式"),
            (88, "CJ", "首年能耗成本偏差", "amount", "报表公式"),
            (89, "CK", "首年回款率偏差", "ratio", "报表公式"),
            (92, "CN", "进场后累计权责税前利润", "amount", "报表公式"),
            (93, "CO", "进场后累计半收付税前利润", "amount", "报表公式"),
            (101, "CW", "进场后累计综合回款率", "ratio", "报表公式"),
        ]

    field_spec_by_index = {item[0]: item for item in field_specs}
    details: list[dict[str, Any]] = []
    source_gaps: Counter[str] = Counter()

    def add_result(
        *,
        row_number_value: int,
        project_code: str,
        column: str,
        field: str,
        kind: str,
        source: str,
        actual: Any,
        expected: Any,
        status: str | None = None,
        alternate: Any = None,
        note: str = "",
    ) -> None:
        passed, difference = compare(actual, expected, kind)
        actual_error = is_excel_error(actual)
        if actual_error:
            passed = False
            difference = None
        final_status = status or (STATUS_PASS if passed else STATUS_FAIL)
        difference_type = ""
        if final_status == STATUS_BLOCKED:
            difference_type = "缺源阻塞"
        elif final_status == STATUS_REVIEW:
            difference_type = "待口径确认"
        elif final_status == STATUS_FAIL:
            expected_empty = expected is None or expected == ""
            actual_number = comparable_number(actual)
            if actual_error:
                difference_type = "报表错误值"
            elif expected_empty and actual_number is not None and near_zero(actual_number):
                difference_type = "应空但报0"
            elif expected_empty:
                difference_type = "应空但有值"
            else:
                difference_type = "公式或取值差异"
        details.append(
            {
                "报表行": row_number_value,
                "项目编码": project_code,
                "列": column,
                "字段": field,
                "报表值": output_value(actual),
                "应有值": output_value(expected),
                "备选口径值": output_value(alternate),
                "差异": output_value(difference),
                "差异类型": difference_type,
                "状态": final_status,
                "来源角色": source,
                "说明": note,
            }
        )

    for excel_row, report_row in enumerate(
        report_sheet.iter_rows(min_row=4, values_only=True), start=4
    ):
        project_code = norm_text(report_row[1] if len(report_row) > 1 else None)
        if not project_code:
            continue

        project_rows = project_map.get(project_code, [])
        project_match_note = "项目编码唯一匹配"
        if not project_rows and args.field_set == "all-report-formulas":
            report_project_name = norm_text(
                report_row[2] if len(report_row) > 2 else None
            )
            name_rows = project_name_map.get(report_project_name, [])
            if len(name_rows) == 1:
                project_rows = name_rows
                project_match_note = "项目编码未命中，按项目名称唯一匹配"
        saturation_rows = saturation_map.get(project_code, [])
        projection_rows = projection_map.get(project_code, [])
        if args.field_set == "external-full":
            saturation_rows = [
                row
                for row in saturation_rows
                if normalized_period(row[2]) == args.report_period
            ]
            projection_rows = [
                row
                for row in projection_rows
                if normalized_period(row[0]) == args.report_period
            ]
        business_rows = business_map.get(project_code, [])
        receivable_rows = receivable_map.get(project_code, [])
        cash_rows = half_cash_map.get(project_code, [])

        project = first_or_none(project_rows)
        if project is None:
            source_gaps["项目主数据缺行或重复"] += 1
            project_expected = {11: None, 12: None}
            if args.field_set != "all-report-formulas":
                project_expected[63] = None
            project_expected[
                9
                if args.field_set
                in {"remaining-report-formulas", "all-report-formulas"}
                else 38
            ] = None
            project_status = STATUS_BLOCKED
        else:
            entry_month = normalized_entry_month(project[12])
            first_year_month = shifted_period(project[12], 11)
            project_expected = {
                11: project[7],
                12: number(project[9]),
            }
            if args.field_set != "all-report-formulas":
                project_expected[63] = (
                    first_year_month
                )
            project_expected[
                9
                if args.field_set
                in {"remaining-report-formulas", "all-report-formulas"}
                else 38
            ] = entry_month
            project_status = None

        for index, column, field, kind, source in field_specs:
            if index not in project_expected:
                continue
            add_result(
                row_number_value=excel_row,
                project_code=project_code,
                column=column,
                field=field,
                kind=kind,
                source=source,
                actual=report_row[index - 1],
                expected=project_expected[index],
                status=project_status,
                note=(
                    project_match_note
                    if project_status is None
                    else "项目主数据无法唯一匹配"
                ),
            )

        saturation_status: str | None = None
        if args.field_set == "external-full" and not saturation_period_available:
            source_gaps["年饱和台账缺报表期间数据"] += 1
            saturation_expected = {13: "", 14: 0.0}
            saturation_status = STATUS_BLOCKED
        elif not saturation_rows:
            source_gaps["年饱和台账项目缺行"] += 1
            saturation_expected = {13: "", 14: 0.0}
            saturation_status = (
                STATUS_BLOCKED if args.missing_row_policy == "blocked" else None
            )
        else:
            years = {year_from_period(row[3]) for row in saturation_rows if row[3]}
            if len(years) != 1:
                source_gaps["年饱和台账年度不唯一"] += 1
                saturation_status = STATUS_BLOCKED
            saturation_expected = {
                13: next(iter(years), ""),
                14: sum_index(saturation_rows, 9),
            }
        for index, column, field, kind, source in field_specs:
            if index not in saturation_expected:
                continue
            add_result(
                row_number_value=excel_row,
                project_code=project_code,
                column=column,
                field=field,
                kind=kind,
                source=source,
                actual=report_row[index - 1],
                expected=saturation_expected[index],
                status=saturation_status,
                note=(
                    f"源台账无报表期间={args.report_period}的数据"
                    if args.field_set == "external-full"
                    and not saturation_period_available
                    else "无匹配行"
                    if not saturation_rows
                    else f"匹配行数={len(saturation_rows)}"
                ),
            )

        projection_status: str | None = None
        projection = first_or_none(projection_rows)
        confirmed_absent = project_code in set(args.confirmed_absent_project_code)
        if (
            args.field_set == "external-full"
            and not projection_period_available
            and not confirmed_absent
        ):
            source_gaps["投模台账缺报表期间数据"] += 1
            projection_status = STATUS_BLOCKED
            projection_values = [0.0] * 15
        elif projection is None:
            reason = "项目缺行" if not projection_rows else "项目重复"
            source_gaps[f"投模台账{reason}"] += 1
            projection_status = STATUS_BLOCKED
            if args.missing_row_policy == "zero" and not projection_rows:
                projection_status = None
            projection_values = [0.0] * 15
        else:
            projection_values = [number(value) for value in projection[3:18]]
        direct_projection = {
            15: projection_values[0],
            16: projection_values[1],
            17: projection_values[2],
            18: projection_values[3],
            19: projection_values[4],
            20: projection_values[5],
            21: projection_values[6],
            22: projection_values[7],
            24: projection_values[9],
            25: projection_values[10],
            26: projection_values[11],
            27: projection_values[12],
            28: projection_values[13],
            29: projection_values[14],
        }
        for index, column, field, kind, source in field_specs:
            if index not in direct_projection:
                continue
            add_result(
                row_number_value=excel_row,
                project_code=project_code,
                column=column,
                field=field,
                kind=kind,
                source=source,
                actual=report_row[index - 1],
                expected=direct_projection[index],
                status=projection_status,
                note=(
                    "用户已确认项目在报表期间投模台账缺行，按0"
                    if confirmed_absent and not projection_rows
                    else f"源台账无报表期间={args.report_period}的数据"
                    if args.field_set == "external-full"
                    and not projection_period_available
                    else "无匹配行"
                    if not projection_rows
                    else f"匹配行数={len(projection_rows)}"
                ),
            )

        # Report-layer formulas are validated using the report's own upstream fields.
        revenue = number(report_row[14])
        semi_revenue = number(report_row[15])
        accrual_profit = number(report_row[16])
        semi_profit = number(report_row[17])
        vat = number(report_row[24])
        management_fee = number(report_row[25])
        monthly_bonus = number(report_row[26])
        financed_input = number(report_row[27])
        adjusted_revenue = revenue
        adjusted_semi_revenue = semi_revenue
        adjusted_accrual_profit = accrual_profit + vat + management_fee + monthly_bonus * 12
        adjusted_semi_profit = semi_profit + vat + management_fee + monthly_bonus * 12
        adjusted_cost = semi_revenue - adjusted_semi_profit
        formula_expected = {
            23: safe_ratio(semi_revenue + revenue * 0.06, revenue * 1.06),
            30: adjusted_revenue,
            31: adjusted_semi_revenue,
            32: adjusted_accrual_profit,
            33: adjusted_semi_profit,
            34: safe_ratio(adjusted_accrual_profit, revenue),
            35: safe_ratio(adjusted_semi_profit, semi_revenue),
            36: adjusted_cost - financed_input,
            37: adjusted_cost,
        }
        if args.field_set in {
            "remaining-report-formulas",
            "all-report-formulas",
        }:
            one_month = shifted_period(report_row[8], 0)
            three_month = shifted_period(report_row[8], 2)
            six_month = shifted_period(report_row[8], 5)
            maturity_enabled = args.field_set == "all-report-formulas"
            one_reached = not maturity_enabled or period_reached(
                one_month, args.report_period
            )
            three_reached = not maturity_enabled or period_reached(
                three_month, args.report_period
            )
            six_reached = not maturity_enabled or period_reached(
                six_month, args.report_period
            )
            formula_expected.update(
                {
                    38: one_month if one_reached else None,
                    42: (
                        safe_ratio(
                            number(report_row[38]),
                            number(report_row[38])
                            + number(report_row[39])
                            - number(report_row[40]),
                        )
                        if one_reached
                        else None
                    ),
                    43: three_month if three_reached else None,
                    47: (
                        safe_ratio(
                            number(report_row[43]),
                            number(report_row[43])
                            + number(report_row[44])
                            - number(report_row[45]),
                        )
                        if three_reached
                        else None
                    ),
                    48: six_month if six_reached else None,
                    51: (
                        number(report_row[48]) - number(report_row[52])
                        if six_reached
                        else None
                    ),
                    52: (
                        number(report_row[49]) - number(report_row[52])
                        if six_reached
                        else None
                    ),
                    60: (
                        safe_ratio(
                            number(report_row[56]),
                            number(report_row[56])
                            + number(report_row[57])
                            - number(report_row[58]),
                        )
                        if six_reached
                        else None
                    ),
                }
            )
        if args.field_set == "all-report-formulas":
            entry_month = shifted_period(report_row[8], 0)
            first_year_month = shifted_period(report_row[8], 11)
            entered = period_reached(entry_month, args.report_period)
            first_year_reached = period_reached(
                first_year_month, args.report_period
            )
            first_year_formula = {
                66: number(report_row[63]) - number(report_row[67]),
                67: number(report_row[64]) - number(report_row[67]),
                75: safe_ratio(
                    number(report_row[71]),
                    number(report_row[71])
                    + number(report_row[72])
                    - number(report_row[73]),
                ),
                78: number(report_row[63]) - number(report_row[29]),
                79: number(report_row[64]) - number(report_row[30]),
                80: number(report_row[65]) - number(report_row[31]),
                81: number(report_row[66]) - number(report_row[32]),
                82: safe_ratio(number(report_row[65]), number(report_row[63]))
                - number(report_row[33]),
                83: safe_ratio(number(report_row[66]), number(report_row[64]))
                - number(report_row[34]),
                84: number(report_row[67]) - number(report_row[36]),
                85: number(report_row[75]) - number(report_row[27]),
                86: number(report_row[68]) - number(report_row[19]),
                87: number(report_row[69]) - number(report_row[20]),
                88: number(report_row[70]) - number(report_row[21]),
                89: number(report_row[74]) - number(report_row[22]),
            }
            formula_expected[63] = (
                first_year_month if first_year_reached else None
            )
            formula_expected.update(
                {
                    index: value if first_year_reached else None
                    for index, value in first_year_formula.items()
                }
            )
            formula_expected.update(
                {
                    92: (
                        number(report_row[89]) - number(report_row[93])
                        if entered
                        else None
                    ),
                    93: (
                        number(report_row[90]) - number(report_row[93])
                        if entered
                        else None
                    ),
                    101: (
                        safe_ratio(
                            number(report_row[97]),
                            number(report_row[97])
                            + number(report_row[98])
                            - number(report_row[99]),
                        )
                        if entered
                        else None
                    ),
                }
            )
        for index, column, field, kind, source in field_specs:
            if index not in formula_expected:
                continue
            add_result(
                row_number_value=excel_row,
                project_code=project_code,
                column=column,
                field=field,
                kind=kind,
                source=source,
                actual=report_row[index - 1],
                expected=formula_expected[index],
                note="按模板报表层公式，以本报表上游字段复算",
            )

        if args.field_set == "external-full":
            accrual_rows = accrual_map.get(project_code, [])
            cumulative_accrual_rows = cumulative_accrual_map.get(project_code, [])
            cumulative_half_rows = cumulative_half_cash_map.get(project_code, [])
            current_accrual_rows = current_accrual_map.get(project_code, [])
            current_half_rows = current_half_cash_map.get(project_code, [])
            current_business_rows = current_business_map.get(project_code, [])
            current_receivable_rows = current_receivable_map.get(project_code, [])
            occupancy_rows = occupancy_map.get(normalized_project_code(project_code), [])
            project_name = norm_text(report_row[2] if len(report_row) > 2 else None)
            amortization_rows = amortization_map.get(project_name, [])

            def add_extended(
                index: int,
                expected: Any,
                note: str,
                status: str | None = None,
            ) -> None:
                spec = field_spec_by_index[index]
                add_result(
                    row_number_value=excel_row,
                    project_code=project_code,
                    column=spec[1],
                    field=spec[2],
                    kind=spec[3],
                    source=spec[4],
                    actual=report_row[index - 1],
                    expected=expected,
                    status=status,
                    note=note,
                )

            for label, rows in (
                ("首年权责底表项目缺行", accrual_rows),
                ("首年半收付底表项目缺行", cash_rows),
                ("首年业务账龄项目缺行", business_rows),
                ("首年应收账龄项目缺行", receivable_rows),
                ("当前业务账龄项目缺行", current_business_rows),
                ("当前应收账龄项目缺行", current_receivable_rows),
            ):
                if not rows:
                    source_gaps[label] += 1

            first_revenue = sum_index(accrual_rows, 11)
            first_semi_revenue = sum_index(cash_rows, 12)
            first_cost = sum_index(accrual_rows, 25)
            first_own_cost = sum_index(accrual_rows, 26) + sum_index(accrual_rows, 27)
            first_outsource_cost = sum_index(accrual_rows, 26)
            first_energy_cost = sum_index(accrual_rows, 29)
            first_cash = sum_index(cash_rows, 10)
            first_receivable = sum_index(business_rows, 11) + sum_index(receivable_rows, 8)
            first_not_due = sum_index(receivable_rows, 14)
            first_direct = {
                64: first_revenue,
                65: first_semi_revenue,
                68: first_cost,
                69: first_own_cost,
                70: first_outsource_cost,
                71: first_energy_cost,
                72: first_cash,
                73: first_receivable,
                74: first_not_due,
            }
            for index, expected in first_direct.items():
                add_extended(
                    index,
                    expected,
                    "首年累计底表；项目缺行按0计算",
                )

            first_formula = {
                66: number(report_row[63]) - number(report_row[67]),
                67: number(report_row[64]) - number(report_row[67]),
                75: safe_ratio(
                    number(report_row[71]),
                    number(report_row[71])
                    + number(report_row[72])
                    - number(report_row[73]),
                ),
            }
            for index, expected in first_formula.items():
                add_extended(index, expected, "按模板公式，以本报表上游字段复算")

            financed_rows = [
                row
                for row in amortization_rows
                if norm_text(row[6] if len(row) > 6 else None) == "带资摊销"
            ]
            financed_input_actual = sum_index(financed_rows, 9)
            if not financed_rows:
                source_gaps["摊销分析无带资摊销行"] += 1
            add_extended(
                76,
                financed_input_actual,
                f"项目名匹配行={len(amortization_rows)}；带资摊销行={len(financed_rows)}；缺行按0",
            )

            occupancy_status: str | None = None
            if not occupancy_rows:
                source_gaps["入住率底表项目缺行"] += 1
                occupancy_expected = 0.0
            elif len(occupancy_rows) == 1:
                occupancy_expected = number(occupancy_rows[0][9])
            else:
                source_gaps["入住率底表项目重复"] += 1
                occupancy_expected = number(occupancy_rows[0][9])
                occupancy_status = STATUS_BLOCKED
            add_extended(
                77,
                occupancy_expected,
                f"项目编码规范化后匹配行={len(occupancy_rows)}；缺行按0",
                occupancy_status,
            )

            deviation_formula = {
                78: number(report_row[63]) - number(report_row[29]),
                79: number(report_row[64]) - number(report_row[30]),
                80: number(report_row[65]) - number(report_row[31]),
                81: number(report_row[66]) - number(report_row[32]),
                82: safe_ratio(number(report_row[65]), number(report_row[63]))
                - number(report_row[33]),
                83: safe_ratio(number(report_row[66]), number(report_row[64]))
                - number(report_row[34]),
                84: number(report_row[67]) - number(report_row[36]),
                85: number(report_row[75]) - number(report_row[27]),
                86: number(report_row[68]) - number(report_row[19]),
                87: number(report_row[69]) - number(report_row[20]),
                88: number(report_row[70]) - number(report_row[21]),
                89: number(report_row[74]) - number(report_row[22]),
            }
            for index, expected in deviation_formula.items():
                add_extended(index, expected, "按模板偏差公式，以本报表上游字段复算")

            if args.cumulative_financial_policy == "sum-period-sources":
                cumulative_revenue = first_revenue + sum_index(
                    current_accrual_rows, 11
                )
                cumulative_semi_revenue = first_semi_revenue + sum_index(
                    current_half_rows, 12
                )
                cumulative_cost = first_cost + sum_index(current_accrual_rows, 25)
                cumulative_own_cost = (
                    first_own_cost
                    + sum_index(current_accrual_rows, 26)
                    + sum_index(current_accrual_rows, 27)
                )
                cumulative_outsource_cost = first_outsource_cost + sum_index(
                    current_accrual_rows, 26
                )
                cumulative_energy_cost = first_energy_cost + sum_index(
                    current_accrual_rows, 29
                )
                cumulative_cash = first_cash + sum_index(current_half_rows, 10)
            else:
                cumulative_revenue = sum_index(cumulative_accrual_rows, 11)
                cumulative_semi_revenue = sum_index(cumulative_half_rows, 12)
                cumulative_cost = sum_index(cumulative_accrual_rows, 25)
                cumulative_own_cost = sum_index(
                    cumulative_accrual_rows, 26
                ) + sum_index(cumulative_accrual_rows, 27)
                cumulative_outsource_cost = sum_index(cumulative_accrual_rows, 26)
                cumulative_energy_cost = sum_index(cumulative_accrual_rows, 29)
                cumulative_cash = sum_index(cumulative_half_rows, 10)
            cumulative_receivable = sum_index(current_business_rows, 11) + sum_index(
                current_receivable_rows, 8
            )
            cumulative_not_due = sum_index(current_receivable_rows, 14)
            cumulative_direct = {
                90: cumulative_revenue,
                91: cumulative_semi_revenue,
                94: cumulative_cost,
                95: cumulative_own_cost,
                96: cumulative_outsource_cost,
                97: cumulative_energy_cost,
                98: cumulative_cash,
                99: cumulative_receivable,
                100: cumulative_not_due,
            }
            for index, expected in cumulative_direct.items():
                status = None
                financial_index = index in {90, 91, 94, 95, 96, 97, 98}
                if financial_index and not cumulative_financial_available:
                    status = STATUS_BLOCKED
                    source_gaps["缺进场月至报表月直接区间财务底表"] += 1
                add_extended(
                    index,
                    expected,
                    (
                        "缺少从进场月至报表月的直接区间权责及半收付底表"
                        if financial_index and not cumulative_financial_available
                        else "累计财务按用户确认口径=首期累计底表+当前期累计底表"
                        if financial_index
                        and args.cumulative_financial_policy == "sum-period-sources"
                        else "使用报表月账龄底表；项目缺行按0"
                        if not financial_index
                        else "使用从进场月至报表月的直接区间底表；项目缺行按0"
                    ),
                    status,
                )

            cumulative_formula = {
                92: number(report_row[89]) - number(report_row[93]),
                93: number(report_row[90]) - number(report_row[93]),
                101: safe_ratio(
                    number(report_row[97]),
                    number(report_row[97])
                    + number(report_row[98])
                    - number(report_row[99]),
                ),
            }
            for index, expected in cumulative_formula.items():
                add_extended(index, expected, "按模板公式，以本报表上游字段复算")
            continue

        if args.field_set in {
            "remaining-report-formulas",
            "all-report-formulas",
        }:
            continue

        receivable_balance = sum_index(business_rows, 11) + sum_index(receivable_rows, 8)
        not_due = sum_index(receivable_rows, 14)
        for index, expected, rows_present, label in (
            (40, receivable_balance, bool(business_rows or receivable_rows), "应收账款余额"),
            (41, not_due, bool(receivable_rows), "未到账期金额"),
        ):
            status = None
            if not rows_present:
                source_gaps[f"账龄底表{label}项目缺行"] += 1
                if args.missing_row_policy == "blocked":
                    status = STATUS_BLOCKED
            spec = next(item for item in field_specs if item[0] == index)
            add_result(
                row_number_value=excel_row,
                project_code=project_code,
                column=spec[1],
                field=spec[2],
                kind=spec[3],
                source=spec[4],
                actual=report_row[index - 1],
                expected=expected,
                status=status,
                note=f"业务账龄行={len(business_rows)}；应收账龄行={len(receivable_rows)}",
            )

        cumulative_cash = sum_index(cash_rows, 10)
        cash_status: str | None = None
        cash_note = f"半收付底表匹配行={len(cash_rows)}；指标清单口径=当年1月至报表月累计"
        if not cash_rows:
            source_gaps["半收付底表项目缺行"] += 1
            if args.missing_row_policy == "blocked":
                cash_status = STATUS_BLOCKED
        window_cash = (
            0.0
            if (cash_rows and near_zero(cumulative_cash))
            or (not cash_rows and args.missing_row_policy == "zero")
            else None
        )
        if project_code in cash_comparative:
            comparative_cumulative, mom, comparative_month = cash_comparative[project_code]
            window_cash = comparative_month
            cash_note += (
                f"；备选进场月口径由累计值及环比反推，累计={comparative_cumulative:.2f}，环比={mom:.12g}"
            )
        actual_cash = report_row[38]
        cumulative_match, _ = compare(actual_cash, cumulative_cash, "amount")
        window_match = False
        if window_cash is not None and not math.isnan(window_cash):
            window_match, _ = compare(actual_cash, window_cash, "amount")
        expected_cash = cumulative_cash
        alternate_cash = window_cash
        if args.cumulative_policy == "entry-month":
            expected_cash = window_cash
            alternate_cash = cumulative_cash
            cash_note += "；采用已确认口径=从进场月开始累计"
            if expected_cash is None or math.isnan(expected_cash):
                cash_status = STATUS_BLOCKED
                source_gaps["缺少进场月累计所需的上期累计值"] += 1
            elif cash_status is None and not window_match:
                cash_status = STATUS_FAIL
        elif args.cumulative_policy == "indicator-year":
            cash_note += "；采用指标清单通用口径=当年1月至报表月累计"
            if cash_status is None and not cumulative_match:
                cash_status = STATUS_FAIL
        elif window_cash is not None and not math.isnan(window_cash):
            if abs(window_cash - cumulative_cash) > AMOUNT_TOLERANCE:
                cash_status = STATUS_REVIEW
                source_gaps["累计现金流时间窗口待确认"] += 1
            elif cash_status is None and not cumulative_match:
                cash_status = STATUS_FAIL
        elif cash_status is None and not cumulative_match:
            cash_status = STATUS_FAIL
        add_result(
            row_number_value=excel_row,
            project_code=project_code,
            column="AM",
            field="累计回收现金流",
            kind="amount",
            source="半收付底表",
            actual=actual_cash,
            expected=expected_cash,
            alternate=alternate_cash,
            status=cash_status,
            note=cash_note
            + f"；累计匹配={cumulative_match}；备选匹配={window_match}",
        )

        # Recovery rate arithmetic is independent of the disputed cash source window.
        report_cash = number(report_row[38])
        report_receivable = number(report_row[39])
        report_not_due = number(report_row[40])
        recovery_expected = safe_ratio(
            report_cash, report_cash + report_receivable - report_not_due
        )
        add_result(
            row_number_value=excel_row,
            project_code=project_code,
            column="AP",
            field="回款率",
            kind="ratio",
            source="报表公式",
            actual=report_row[41],
            expected=recovery_expected,
            note="按模板公式，以本报表AM、AN、AO复算；源值结论依赖AM口径确认",
        )

    if args.field_set in {
        "remaining-report-formulas",
        "all-report-formulas",
    }:
        # This field set deliberately validates only project-master and
        # report-layer relationships. Do not surface unused ledger gaps.
        source_gaps = Counter(
            {
                key: value
                for key, value in source_gaps.items()
                if key.startswith("项目主数据")
            }
        )

    counts = Counter(item["状态"] for item in details)
    per_field: dict[str, Counter[str]] = defaultdict(Counter)
    per_field_difference: dict[str, Counter[str]] = defaultdict(Counter)
    for item in details:
        field_key = f"{item['列']} {item['字段']}"
        per_field[field_key][item["状态"]] += 1
        if item["差异类型"]:
            per_field_difference[field_key][item["差异类型"]] += 1

    summary = {
        "report_rows": len({item["报表行"] for item in details}),
        "field_count": len(field_specs),
        "detail_count": len(details),
        "status_counts": dict(counts),
        "source_gaps": dict(source_gaps),
        "field_set": args.field_set,
        "report_period": args.report_period,
        "missing_row_policy": args.missing_row_policy,
        "cumulative_policy": args.cumulative_policy,
        "cumulative_financial_policy": args.cumulative_financial_policy,
        "template_evidence": template_check,
        "indicator_evidence": indicator_check,
        "field_status": {key: dict(value) for key, value in per_field.items()},
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "验证汇总"
    summary_rows = [
        ("项目", "值"),
        ("报表项目数", summary["report_rows"]),
        ("字段数", summary["field_count"]),
        ("明细检查数", summary["detail_count"]),
        ("通过", counts.get(STATUS_PASS, 0)),
        ("失败", counts.get(STATUS_FAIL, 0)),
        ("缺源阻塞", counts.get(STATUS_BLOCKED, 0)),
        ("待口径确认", counts.get(STATUS_REVIEW, 0)),
        ("缺行政策", args.missing_row_policy),
        ("字段集", args.field_set),
        ("报表期间", args.report_period or ""),
        ("累计政策", args.cumulative_policy),
        ("累计财务来源政策", args.cumulative_financial_policy),
        ("说明", "敏感数据仅保存在本地输出；状态按逐项目逐字段统计。"),
    ]
    for row in summary_rows:
        summary_sheet.append(row)
    summary_sheet.append([])
    summary_sheet.append(["字段", STATUS_PASS, STATUS_FAIL, STATUS_BLOCKED, STATUS_REVIEW])
    for field in sorted(per_field, key=lambda value: field_specs[[x[1] for x in field_specs].index(value.split()[0])][0]):
        field_counts = per_field[field]
        summary_sheet.append(
            [
                field,
                field_counts.get(STATUS_PASS, 0),
                field_counts.get(STATUS_FAIL, 0),
                field_counts.get(STATUS_BLOCKED, 0),
                field_counts.get(STATUS_REVIEW, 0),
            ]
        )

    details_sheet = workbook.create_sheet("逐行明细")
    detail_headers = list(details[0].keys()) if details else []
    details_sheet.append(detail_headers)
    for item in details:
        details_sheet.append([item[header] for header in detail_headers])

    gaps_sheet = workbook.create_sheet("缺源与待确认")
    gaps_sheet.append(["事项", "项目字段次数"])
    for key, value in source_gaps.items():
        gaps_sheet.append([key, value])

    difference_sheet = workbook.create_sheet("差异归因汇总")
    difference_headers = [
        "字段",
        "应空但报0",
        "应空但有值",
        "公式或取值差异",
        "报表错误值",
        "缺源阻塞",
        "待口径确认",
    ]
    difference_sheet.append(difference_headers)
    for field in sorted(
        per_field,
        key=lambda value: field_specs[
            [x[1] for x in field_specs].index(value.split()[0])
        ][0],
    ):
        field_differences = per_field_difference[field]
        difference_sheet.append(
            [field]
            + [field_differences.get(header, 0) for header in difference_headers[1:]]
        )

    if args.field_set in {
        "remaining-report-formulas",
        "all-report-formulas",
    }:
        basis_sheet = workbook.create_sheet("字段识别依据")
        basis_sheet.append(
            ["列", "字段", "计算关系", "模板第5行", "纳入原因"]
        )
        basis_rows = [
            ("I", "进场时间", "进场日期1-14日计当月，15日后计次月", "非报表层", "存在换月计算"),
            ("AL", "1个月时间", "等于进场时间", "yyyy-mm", "模板第3行明确等于关系"),
            ("AP", "1个月回款率", "现金流/(现金流+应收-未到账期)", "报表层", "明确报表公式"),
            ("AQ", "3个月时间", "进场时间+2个月", "空", "模板第3行明确日期偏移"),
            ("AU", "3个月回款率", "现金流/(现金流+应收-未到账期)", "报表层", "明确报表公式"),
            ("AV", "6个月时间", "进场时间+5个月", "空", "模板第3行明确日期偏移"),
            ("AY", "6个月权责税前利润", "营业收入-营业成本", "空", "模板第3行明确算式"),
            ("AZ", "6个月半收付税前利润", "半收付收入-营业成本", "空", "模板第3行明确算式"),
            ("BH", "6个月回款率", "现金流/(现金流+应收-未到账期)", "报表层", "明确报表公式"),
        ]
        if args.field_set == "all-report-formulas":
            formulas = {
                "I": "项目进场日期按1-14日当月、15日后次月换月",
                "W": "(P+O*6%)/(O*106%)",
                "AD": "O",
                "AE": "P",
                "AF": "Q+Y+Z+AA*12",
                "AG": "R+Y+Z+AA*12",
                "AH": "AF/O",
                "AI": "AG/P",
                "AJ": "AK-AB",
                "AK": "P-AG",
                "AL": "I；未满1个月为空",
                "AP": "AM/(AM+AN-AO)；未满1个月为空",
                "AQ": "I+2个月；未满3个月为空",
                "AU": "AR/(AR+AS-AT)；未满3个月为空",
                "AV": "I+5个月；未满6个月为空",
                "AY": "AW-BA；未满6个月为空",
                "AZ": "AX-BA；未满6个月为空",
                "BH": "BE/(BE+BF-BG)；未满6个月为空",
                "BK": "I+11个月；未满12个月为空",
                "BN": "BL-BP；未满12个月为空",
                "BO": "BM-BP；未满12个月为空",
                "BW": "BT/(BT+BU-BV)；未满12个月为空",
                "BZ": "BL-AD；未满12个月为空",
                "CA": "BM-AE；未满12个月为空",
                "CB": "BN-AF；未满12个月为空",
                "CC": "BO-AG；未满12个月为空",
                "CD": "BN/BL-AH；未满12个月为空",
                "CE": "BO/BM-AI；未满12个月为空",
                "CF": "BP-AK；未满12个月为空",
                "CG": "BX-AB；未满12个月为空",
                "CH": "BQ-T；未满12个月为空",
                "CI": "BR-U；未满12个月为空",
                "CJ": "BS-V；未满12个月为空",
                "CK": "BW-W；未满12个月为空",
                "CN": "CL-CP；未进场为空",
                "CO": "CM-CP；未进场为空",
                "CW": "CT/(CT+CU-CV)；未进场为空",
            }
            basis_rows = [
                (
                    column,
                    field,
                    formulas[column],
                    "不作为唯一依据",
                    "模板计算关系或实际同表派生",
                )
                for _, column, field, _, _ in field_specs
            ]
        for row in basis_rows:
            basis_sheet.append(row)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    status_fills = {
        STATUS_PASS: PatternFill("solid", fgColor="E2F0D9"),
        STATUS_FAIL: PatternFill("solid", fgColor="F4CCCC"),
        STATUS_BLOCKED: PatternFill("solid", fgColor="FCE5CD"),
        STATUS_REVIEW: PatternFill("solid", fgColor="FFF2CC"),
    }
    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2"
        sheet.sheet_view.showGridLines = False
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column_cells in sheet.columns:
            width = min(
                60,
                max(10, max(len(norm_text(cell.value)) for cell in column_cells) + 2),
            )
            sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = width
    gaps_sheet.column_dimensions["B"].width = 18
    summary_sheet.column_dimensions["A"].width = 42
    summary_sheet.column_dimensions["B"].width = 22
    for column in ("C", "D", "E"):
        summary_sheet.column_dimensions[column].width = 16
    difference_sheet.column_dimensions["A"].width = 42
    for column in ("B", "C", "D", "E", "F", "G"):
        difference_sheet.column_dimensions[column].width = 18
    details_sheet.column_dimensions["D"].width = 34
    details_sheet.column_dimensions["I"].width = 18
    details_sheet.column_dimensions["L"].width = 52
    if args.field_set in {
        "remaining-report-formulas",
        "all-report-formulas",
    }:
        basis_sheet.column_dimensions["B"].width = 34
        basis_sheet.column_dimensions["C"].width = 48
        basis_sheet.column_dimensions["D"].width = 20
        basis_sheet.column_dimensions["E"].width = 30
    if details:
        status_column = detail_headers.index("状态") + 1
        for row_index in range(2, details_sheet.max_row + 1):
            cell = details_sheet.cell(row_index, status_column)
            if cell.value in status_fills:
                cell.fill = status_fills[cell.value]
        details_sheet.auto_filter.ref = details_sheet.dimensions
    summary_sheet.auto_filter.ref = f"A{len(summary_rows)+2}:E{summary_sheet.max_row}"
    workbook.save(args.output)

    if args.summary_json:
        args.summary_json.parent.mkdir(parents=True, exist_ok=True)
        args.summary_json.write_text(
            json.dumps(summary, ensure_ascii=False, indent=2), encoding="utf-8"
        )

    print(json.dumps(summary, ensure_ascii=False, indent=2))
    return 0 if counts.get(STATUS_FAIL, 0) == 0 else 2


if __name__ == "__main__":
    raise SystemExit(main())
