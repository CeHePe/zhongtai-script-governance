from __future__ import annotations

import argparse
import sys
from collections import defaultdict
from dataclasses import dataclass
from pathlib import Path
from typing import Any, Iterable

from openpyxl import load_workbook

from _project_root import find_project_root


REPORT_MONTH = "2025-12"
PREVIOUS_REPORT_MONTH = "2024-12"
ROUND_TOLERANCE = 0.0001
ZERO_TOLERANCE_WAN = 0.000001


COL_REPORT_MONTH = "报表年月"
COL_PERFORMANCE_MONTH = "业绩认定年月"
COL_BUSINESS = "大业态"
COL_TYPE = "类型"
COL_REGION = "区域"
COL_BEFORE = "年度饱和收入_打折前（元）"
COL_AFTER = "口径认定金额（元）"

VALUE_RESIDENTIAL = "住宅"
VALUE_JINYI = "金颐"
VALUE_SUMMARY = "汇总"


@dataclass(frozen=True)
class DimensionConfig:
    code: str
    name: str
    report_name_part: str
    indicator_dimension: str
    is_region_report: bool


DIMENSIONS = [
    DimensionConfig(
        code="region_residential",
        name="区域住宅",
        report_name_part="区域住宅",
        indicator_dimension="区域住宅",
        is_region_report=True,
    ),
    DimensionConfig(
        code="region_enterprise",
        name="区域政企",
        report_name_part="区域政企",
        indicator_dimension="区域政企",
        is_region_report=True,
    ),
    DimensionConfig(
        code="jinyi",
        name="金颐",
        report_name_part="金颐",
        indicator_dimension="金颐",
        is_region_report=False,
    ),
]


def configure_stdout() -> None:
    if hasattr(sys.stdout, "reconfigure"):
        sys.stdout.reconfigure(encoding="utf-8")


def find_workbook(base: Path, predicate, description: str) -> Path:
    matches = [
        path
        for path in base.iterdir()
        if path.is_file()
        and path.suffix.lower() == ".xlsx"
        and not path.name.startswith("~$")
        and predicate(path.name)
    ]
    if not matches:
        raise FileNotFoundError(f"未找到{description}")
    matches.sort(key=lambda path: path.name)
    return matches[0]


def workbook_rows(path: Path) -> list[tuple[Any, ...]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook.active
    # 部分系统导出的 xlsx 维度标记只有 A1，必须重置后才能读出真实区域。
    worksheet.reset_dimensions()
    return list(worksheet.iter_rows(values_only=True))


def table_rows(path: Path) -> list[dict[str, Any]]:
    rows = workbook_rows(path)
    if not rows:
        return []
    headers = [str(value).strip() if value is not None else "" for value in rows[0]]
    records: list[dict[str, Any]] = []
    for row in rows[1:]:
        if all(value in (None, "") for value in row):
            continue
        records.append({header: row[index] if index < len(row) else None for index, header in enumerate(headers)})
    return records


def normalized_month(value: Any) -> str:
    if value is None:
        return ""
    return str(value)[:7]


def as_float(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    return float(value)


def calc_yoy(numerator: float, denominator: float) -> float:
    if abs(denominator) < ZERO_TOLERANCE_WAN:
        return 0.0
    return numerator / denominator - 1


def close_enough(actual: float, expected: float) -> bool:
    return abs(actual - expected) <= ROUND_TOLERANCE


def row_matches_dimension(row: dict[str, Any], config: DimensionConfig, *, require_region: bool) -> bool:
    business = row.get(COL_BUSINESS)
    type_value = row.get(COL_TYPE)
    region = row.get(COL_REGION)
    if config.code == "region_residential":
        return business == VALUE_RESIDENTIAL and (not require_region or region not in (None, ""))
    if config.code == "region_enterprise":
        return business != VALUE_RESIDENTIAL and type_value != VALUE_JINYI and (
            not require_region or region not in (None, "")
        )
    if config.code == "jinyi":
        return type_value == VALUE_JINYI
    raise ValueError(f"未知维度: {config.code}")


def aggregate_current(records: Iterable[dict[str, Any]], config: DimensionConfig) -> dict[str, dict[str, float]]:
    aggregates: dict[str, dict[str, float]] = defaultdict(
        lambda: {"after": 0.0, "before": 0.0, "n_minus_1_after": 0.0, "n_after": 0.0, "count": 0.0}
    )
    for row in records:
        if normalized_month(row.get(COL_REPORT_MONTH)) != REPORT_MONTH:
            continue
        if not row_matches_dimension(row, config, require_region=config.is_region_report):
            continue

        key = str(row.get(COL_REGION)) if config.is_region_report else config.name
        after_wan = as_float(row.get(COL_AFTER)) / 10000
        before_wan = as_float(row.get(COL_BEFORE)) / 10000
        aggregates[key]["after"] += after_wan
        aggregates[key]["before"] += before_wan
        aggregates[key]["count"] += 1
        if normalized_month(row.get(COL_PERFORMANCE_MONTH)) == REPORT_MONTH:
            aggregates[key]["n_after"] += after_wan
        else:
            aggregates[key]["n_minus_1_after"] += after_wan
    return dict(aggregates)


def aggregate_previous(records: Iterable[dict[str, Any]], config: DimensionConfig) -> dict[str, float]:
    aggregates: dict[str, float] = defaultdict(float)
    for row in records:
        if normalized_month(row.get(COL_REPORT_MONTH)) != PREVIOUS_REPORT_MONTH:
            continue
        if not row_matches_dimension(row, config, require_region=config.is_region_report):
            continue
        key = str(row.get(COL_REGION)) if config.is_region_report else config.name
        aggregates[key] += as_float(row.get(COL_BEFORE)) / 10000
    return dict(aggregates)


def read_report(path: Path, config: DimensionConfig) -> tuple[dict[str, dict[str, float]], dict[str, float]]:
    rows = workbook_rows(path)
    if config.is_region_report:
        details: dict[str, dict[str, float]] = {}
        summary: dict[str, float] | None = None
        for row in rows[2:]:
            first = row[0] if len(row) > 0 else None
            if first in (None, ""):
                continue
            values = {
                "after": as_float(row[2]),
                "target": as_float(row[3]),
                "target_rate": as_float(row[4]),
                "yoy": as_float(row[5]),
                "target_gap": as_float(row[6]),
                "n_minus_1_after": as_float(row[7]),
                "n_after": as_float(row[8]),
            }
            if first == VALUE_SUMMARY:
                summary = values
            else:
                details[str(first)] = values
        if summary is None:
            raise RuntimeError(f"{path.name} 缺少汇总行")
        return details, summary

    numeric_rows = []
    for row in rows[2:]:
        if row and row[0] not in (None, ""):
            numeric_rows.append(
                {
                    "after": as_float(row[0]),
                    "target": as_float(row[1]),
                    "target_rate": as_float(row[2]),
                    "yoy": as_float(row[3]),
                    "target_gap": as_float(row[4]),
                    "n_minus_1_after": as_float(row[5]),
                    "n_after": as_float(row[6]),
                }
            )
    if not numeric_rows:
        raise RuntimeError(f"{path.name} 缺少数据行")
    return {config.name: numeric_rows[0]}, numeric_rows[-1]


def read_indicator_rows(base: Path) -> list[dict[str, Any]]:
    indicator_path = find_workbook(
        base,
        lambda name: name.startswith("JKS_") and name.endswith("指标清单.xlsx"),
        "指标清单",
    )
    workbook = load_workbook(indicator_path, read_only=True, data_only=True)
    worksheet = workbook.active
    headers = [cell.value if cell.value is not None else f"未命名列{index}" for index, cell in enumerate(next(worksheet.iter_rows()), 1)]
    rows: list[dict[str, Any]] = []
    target_dimensions = {config.indicator_dimension for config in DIMENSIONS}
    for excel_row, row in enumerate(worksheet.iter_rows(min_row=2, values_only=True), start=2):
        row_dict = {str(headers[index]).strip(): value for index, value in enumerate(row)}
        metric_name = str(row_dict.get("指标名称") or row_dict.get("指标名称".rjust(66)) or "")
        dimension = row_dict.get("组织维度")
        logic = str(row_dict.get("计算逻辑") or "")
        if dimension in target_dimensions and "新增年度饱和收入台账" in logic and (
            "年度饱和收入" in metric_name or "口径认定金额" in logic
        ):
            rows.append({"excel_row": excel_row, **row_dict})
    return rows


def validate_dimension(base: Path, current_rows: list[dict[str, Any]], previous_rows: list[dict[str, Any]], config: DimensionConfig) -> list[dict[str, Any]]:
    report_path = find_workbook(
        base,
        lambda name: name.startswith("1.4.1_") and REPORT_MONTH.replace("-", "") in name and config.report_name_part in name,
        f"1.4.1 {config.name}报表",
    )
    report_details, report_summary = read_report(report_path, config)
    current = aggregate_current(current_rows, config)
    previous = aggregate_previous(previous_rows, config)

    results: list[dict[str, Any]] = []
    for key in sorted(report_details):
        expected = current.get(key, {"after": 0.0, "before": 0.0, "n_minus_1_after": 0.0, "n_after": 0.0})
        previous_before = previous.get(key, 0.0)
        expected_yoy = calc_yoy(expected["before"], previous_before)
        report_row = report_details[key]
        results.append(
            {
                "dimension": config.name,
                "row": key,
                "kind": "明细",
                "report_after": report_row["after"],
                "expected_after": expected["after"],
                "after_pass": close_enough(report_row["after"], expected["after"]),
                "report_yoy": report_row["yoy"],
                "expected_yoy": expected_yoy,
                "yoy_pass": close_enough(report_row["yoy"], expected_yoy),
                "current_before": expected["before"],
                "previous_before": previous_before,
            }
        )

    total_after = sum(value["after"] for value in current.values())
    total_before = sum(value["before"] for value in current.values())
    total_previous_before = sum(previous.values())
    expected_summary_yoy = calc_yoy(total_before, total_previous_before)
    results.append(
        {
            "dimension": config.name,
            "row": VALUE_SUMMARY,
            "kind": "汇总",
            "report_after": report_summary["after"],
            "expected_after": total_after,
            "after_pass": close_enough(report_summary["after"], total_after),
            "report_yoy": report_summary["yoy"],
            "expected_yoy": expected_summary_yoy,
            "yoy_pass": close_enough(report_summary["yoy"], expected_summary_yoy),
            "current_before": total_before,
            "previous_before": total_previous_before,
        }
    )
    return results


def print_indicator_confirmation(rows: list[dict[str, Any]]) -> None:
    print("指标清单确认：")
    for row in rows:
        metric_name = row.get("指标名称") or row.get("指标名称".rjust(66))
        print(
            f"- Excel行{row['excel_row']} 序号={row.get('序号')} "
            f"指标={metric_name} 维度={row.get('组织维度')} "
            f"累计/月度={row.get('累计/月度')} 取数表={row.get('取数对应表')}"
        )
    print("- D类已撤场排除：指标清单未要求")
    print("- （利润）非考核项目排除：指标清单未要求")
    print("- 高维附加调整：指标清单未要求")
    print("- 高维汇总方式：本指标直接从新增年度饱和收入台账按报表年月与维度汇总，不依赖项目维报表")


def print_results(results: list[dict[str, Any]]) -> None:
    print("\n验证结果：")
    print("维度\t行\t类型\t打折后报表\t打折后应为\t打折后结论\t同比报表\t同比应为\t同比结论\t汇总分子\t汇总分母")
    for result in results:
        after_status = "PASS" if result["after_pass"] else "FAIL"
        yoy_status = "PASS" if result["yoy_pass"] else "FAIL"
        print(
            f"{result['dimension']}\t{result['row']}\t{result['kind']}\t"
            f"{result['report_after']:.6f}\t{result['expected_after']:.6f}\t{after_status}\t"
            f"{result['report_yoy']:.6f}\t{result['expected_yoy']:.6f}\t{yoy_status}\t"
            f"{result['current_before']:.6f}\t{result['previous_before']:.6f}"
        )
    failures = [result for result in results if not (result["after_pass"] and result["yoy_pass"])]
    print(f"\n汇总：共{len(results)}行，失败{len(failures)}行")


def main() -> None:
    configure_stdout()
    parser = argparse.ArgumentParser(description="验证1.4.1直拓经营分析的年度饱和收入与同比。")
    parser.parse_args()

    base = find_project_root(__file__)
    current_path = find_workbook(base, lambda name: name.startswith("新增年度饱和收入台账") and "202512" in name, "202512新增年度饱和收入台账")
    previous_path = find_workbook(base, lambda name: name.startswith("新增年度饱和收入台账") and "202412" in name, "202412新增年度饱和收入台账")
    current_rows = table_rows(current_path)
    previous_rows = table_rows(previous_path)

    print(f"本地来源：{current_path.name}（{len(current_rows)}行）、{previous_path.name}（{len(previous_rows)}行）")
    print_indicator_confirmation(read_indicator_rows(base))

    all_results: list[dict[str, Any]] = []
    for config in DIMENSIONS:
        all_results.extend(validate_dimension(base, current_rows, previous_rows, config))
    print_results(all_results)


if __name__ == "__main__":
    main()
