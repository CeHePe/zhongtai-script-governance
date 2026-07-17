from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


TOLERANCE_YUAN = 0.01
FUNDED_TYPE_CODE = "TX01"
YES_VALUES = {"是", "1", "true", "mq=="}


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description=(
            "Validate BI/BX funded-input windows in the external-quality project report. "
            "Finance-cloud A is assigned to the approval month; business-finance B is zero "
            "when the locally confirmed source has no funded rows."
        )
    )
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--project-query", type=Path, required=True)
    parser.add_argument("--finance-history", type=Path, required=True)
    parser.add_argument("--finance-supplement", type=Path, required=True)
    parser.add_argument(
        "--business-finance-policy",
        choices=["confirmed-zero"],
        required=True,
        help="Explicit confirmation that business-finance B has no funded rows.",
    )
    parser.add_argument("--report-period", required=True, help="YYYYMM")
    parser.add_argument("--output-json", type=Path, required=True)
    return parser.parse_args()


def text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def number_or_none(value: Any) -> float | None:
    if value in (None, ""):
        return None
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return None if pd.isna(number) else number


def is_blank(value: Any) -> bool:
    return value is None or text(value) == ""


def is_blank_or_zero(value: Any) -> bool:
    if is_blank(value):
        return True
    number = number_or_none(value)
    return number is not None and abs(number) < TOLERANCE_YUAN


def period(value: Any) -> pd.Period | None:
    digits = re.sub(r"\D", "", text(value))
    if len(digits) < 6:
        return None
    try:
        return pd.Period(f"{digits[:4]}-{digits[4:6]}", freq="M")
    except ValueError:
        return None


def funded_mask(frame: pd.DataFrame, type_code: str, type_name: str) -> pd.Series:
    return frame[type_code].astype(str).str.strip().eq(FUNDED_TYPE_CODE) | frame[
        type_name
    ].astype(str).str.contains("带资", na=False)


def load_finance_a(
    history_path: Path, supplement_path: Path
) -> tuple[pd.DataFrame, list[dict[str, Any]]]:
    history = pd.read_excel(history_path)
    required_history = {
        "is_shared",
        "share_type_code",
        "share_type_name",
        "sub_project_code",
        "project_code",
        "no_tax_dist_amount",
        "approval_time",
    }
    missing = required_history - set(history.columns)
    if missing:
        raise ValueError(f"finance history missing columns: {sorted(missing)}")
    shared = history["is_shared"].astype(str).str.strip().str.lower().isin(YES_VALUES)
    history = history[shared & funded_mask(history, "share_type_code", "share_type_name")].copy()
    history["项目编码"] = history["sub_project_code"].fillna(history["project_code"]).map(text)
    history["审批通过时间"] = pd.to_datetime(history["approval_time"], errors="coerce")
    history["分摊不含税金额"] = pd.to_numeric(
        history["no_tax_dist_amount"], errors="coerce"
    )
    history["来源"] = "财务云历史实际"

    supplement = pd.read_excel(supplement_path)
    required_supplement = {
        "是否分摊",
        "摊销类型编码",
        "摊销类型名称",
        "立项编码",
        "分摊不含税金额",
        "审批通过时间",
    }
    missing = required_supplement - set(supplement.columns)
    if missing:
        raise ValueError(f"finance supplement missing columns: {sorted(missing)}")
    shared = supplement["是否分摊"].astype(str).str.strip().eq("是")
    supplement = supplement[
        shared & funded_mask(supplement, "摊销类型编码", "摊销类型名称")
    ].copy()
    supplement["项目编码"] = supplement["立项编码"].map(text)
    supplement["审批通过时间"] = pd.to_datetime(
        supplement["审批通过时间"], errors="coerce"
    )
    supplement["分摊不含税金额"] = pd.to_numeric(
        supplement["分摊不含税金额"], errors="coerce"
    )
    supplement["来源"] = "财务云补充实际"

    actual = pd.concat(
        [
            history[["项目编码", "审批通过时间", "分摊不含税金额", "来源"]],
            supplement[["项目编码", "审批通过时间", "分摊不含税金额", "来源"]],
        ],
        ignore_index=True,
    )
    invalid = actual[
        actual["项目编码"].eq("")
        | actual["审批通过时间"].isna()
        | actual["分摊不含税金额"].isna()
    ]
    if not invalid.empty:
        raise ValueError(f"funded finance A has {len(invalid)} rows missing code/date/amount")
    actual["月份"] = actual["审批通过时间"].dt.to_period("M")

    audits: list[dict[str, Any]] = []
    for (source, approval_date), group in actual.groupby(
        ["来源", actual["审批通过时间"].dt.strftime("%Y-%m-%d")], dropna=False
    ):
        audits.append(
            {
                "来源": source,
                "审批通过日期": approval_date,
                "带资行数": len(group),
                "项目编码数": group["项目编码"].nunique(),
                "分摊不含税金额": float(group["分摊不含税金额"].sum()),
            }
        )
    return actual, audits


def load_penetration(path: Path) -> tuple[dict[str, float], set[str]]:
    frame = pd.read_excel(path)
    required = {"立项编码", "穿透比例"}
    missing = required - set(frame.columns)
    if missing:
        raise ValueError(f"project query missing columns: {sorted(missing)}")
    frame["立项编码"] = frame["立项编码"].map(text)
    frame["穿透比例"] = pd.to_numeric(frame["穿透比例"], errors="coerce")
    ratios: dict[str, float] = {}
    conflicting: set[str] = set()
    for code, group in frame[frame["立项编码"].ne("")].groupby("立项编码"):
        values = sorted({float(value) for value in group["穿透比例"].dropna()})
        if len(values) == 1:
            ratios[code] = values[0]
        elif len(values) > 1:
            conflicting.add(code)
    return ratios, conflicting


def iter_report_rows(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    sheet.reset_dimensions()
    yield from enumerate(sheet.iter_rows(min_row=4, values_only=True), 4)


def validate(args: argparse.Namespace) -> dict[str, Any]:
    report_period = period(args.report_period)
    if report_period is None or re.sub(r"\D", "", args.report_period)[:6] != args.report_period:
        raise ValueError("--report-period must use YYYYMM")
    for path in (
        args.report,
        args.project_query,
        args.finance_history,
        args.finance_supplement,
    ):
        if not path.exists():
            raise FileNotFoundError(path)

    finance_a, source_audits = load_finance_a(
        args.finance_history, args.finance_supplement
    )
    finance_by_code = {
        code: group.copy() for code, group in finance_a.groupby("项目编码")
    }
    penetration, conflicting_ratios = load_penetration(args.project_query)

    field_specs = [
        {"字段列": "BI", "字段": "投后6个月带资投入费用", "列索引": 60, "月数": 6},
        {"字段列": "BX", "字段": "投后首年带资投入费用", "列索引": 75, "月数": 12},
    ]
    details: list[dict[str, Any]] = []
    report_rows = 0
    for excel_row, row in iter_report_rows(args.report):
        report_rows += 1
        code = text(row[1] if len(row) > 1 else None)
        project_name = text(row[2] if len(row) > 2 else None)
        entry = period(row[8] if len(row) > 8 else None)
        report_ratio = number_or_none(row[11] if len(row) > 11 else None)
        ratio = penetration.get(code)
        code_actual = finance_by_code.get(code)

        for spec in field_specs:
            raw_report_value = row[spec["列索引"]] if len(row) > spec["列索引"] else None
            window_end = entry + (spec["月数"] - 1) if entry is not None else None
            mature = window_end is not None and window_end <= report_period
            source_rows = pd.DataFrame()
            raw_sum: float | None = None
            expected_a: float | None = None
            status = ""
            difference_type = ""
            explanation = ""

            if entry is None:
                status = "通过" if is_blank_or_zero(raw_report_value) else "失败"
                difference_type = "" if status == "通过" else "未成熟非零"
                explanation = "进场月缺失，无法形成投后窗口；空值允许以0体现"
            elif code in conflicting_ratios:
                status = "阻塞"
                difference_type = "穿透比例冲突"
                explanation = "项目查询中同一立项编码存在多个穿透比例"
            elif ratio is None:
                status = "阻塞"
                difference_type = "穿透比例缺失"
                explanation = "项目查询中没有可用穿透比例"
            elif not mature:
                status = "通过" if is_blank_or_zero(raw_report_value) else "失败"
                difference_type = "" if status == "通过" else "未成熟非零"
                explanation = f"未满{spec['月数']}个月；空值允许以0体现"
            else:
                if code_actual is not None:
                    source_rows = code_actual[
                        code_actual["月份"].between(entry, window_end)
                    ].copy()
                raw_sum = (
                    0.0
                    if source_rows.empty
                    else float(source_rows["分摊不含税金额"].sum())
                )
                expected_a = raw_sum * 0.81 * ratio
                expected = expected_a  # B path is locally confirmed absent and therefore zero.
                report_value = number_or_none(raw_report_value)
                actual = 0.0 if report_value is None else report_value
                diff = actual - expected
                status = "通过" if abs(diff) < TOLERANCE_YUAN else "失败"
                difference_type = "" if status == "通过" else "金额差异"
                explanation = (
                    "A按审批通过时间落单月并汇总窗口；B经用户确认无带资记录按0；"
                    "成熟字段缺值按0"
                )

            report_value = number_or_none(raw_report_value)
            expected_total = expected_a if mature and expected_a is not None else None
            diff = (
                (0.0 if report_value is None else report_value) - expected_total
                if expected_total is not None
                else None
            )
            trace = ""
            if not source_rows.empty:
                trace = "；".join(
                    f"{item['月份']}:{float(item['分摊不含税金额']):.6f}({item['来源']})"
                    for _, item in source_rows.sort_values(["月份", "来源"]).iterrows()
                )
            details.append(
                {
                    "报表行": excel_row,
                    "项目编码": code,
                    "项目名称": project_name,
                    "字段列": spec["字段列"],
                    "字段": spec["字段"],
                    "进场月": "" if entry is None else str(entry),
                    "窗口结束月": "" if window_end is None else str(window_end),
                    "是否成熟": "是" if mature else "否",
                    "报表原始值": raw_report_value,
                    "报表值": report_value,
                    "报表股权比例": report_ratio,
                    "项目查询穿透比例": ratio,
                    "A原始金额合计": raw_sum,
                    "A调整后金额": expected_a,
                    "B金额": 0.0 if mature else None,
                    "期望值": expected_total,
                    "差异": diff,
                    "A命中行数": len(source_rows),
                    "A明细": trace,
                    "差异类型": difference_type,
                    "状态": status,
                    "说明": explanation,
                }
            )

    summary: list[dict[str, Any]] = []
    for spec in field_specs:
        rows = [item for item in details if item["字段列"] == spec["字段列"]]
        mature_rows = [item for item in rows if item["是否成熟"] == "是"]
        immature_rows = [item for item in rows if item["是否成熟"] == "否"]
        counts = Counter(item["状态"] for item in rows)
        summary.append(
            {
                "字段列": spec["字段列"],
                "字段": spec["字段"],
                "总项目": len(rows),
                "成熟项目": len(mature_rows),
                "未成熟项目": len(immature_rows),
                "成熟A命中项目": sum(item["A命中行数"] > 0 for item in mature_rows),
                "成熟报表非零项目": sum(
                    abs(item["报表值"] or 0.0) >= TOLERANCE_YUAN
                    for item in mature_rows
                ),
                "通过": counts.get("通过", 0),
                "失败": counts.get("失败", 0),
                "阻塞": counts.get("阻塞", 0),
                "金额差异": sum(item["差异类型"] == "金额差异" for item in rows),
                "未成熟非零": sum(
                    item["差异类型"] == "未成熟非零" for item in rows
                ),
            }
        )

    return {
        "metadata": {
            "report_period": args.report_period,
            "tolerance_yuan": TOLERANCE_YUAN,
            "finance_a_rule": "是否分摊=是；带资；按审批通过时间作为单月归属；分摊不含税金额*0.81*穿透比例",
            "finance_b_rule": "用户确认业财无带资记录，按0",
            "business_finance_policy": args.business_finance_policy,
            "maturity_rule": (
                "BI未满6个月、BX未满12个月时为空，报表以0体现也视为通过；"
                "成熟字段缺值按0"
            ),
            "report_rows": report_rows,
            "finance_a_rows": len(finance_a),
            "finance_a_codes": finance_a["项目编码"].nunique(),
            "finance_a_min_date": finance_a["审批通过时间"].min().strftime("%Y-%m-%d"),
            "finance_a_max_date": finance_a["审批通过时间"].max().strftime("%Y-%m-%d"),
            "conflicting_penetration_codes": len(conflicting_ratios),
        },
        "summary": summary,
        "source_audits": source_audits,
        "details": details,
    }


def main() -> None:
    args = parse_args()
    result = validate(args)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(result, ensure_ascii=False, indent=2, default=str), encoding="utf-8"
    )
    print(json.dumps({"metadata": result["metadata"], "summary": result["summary"]}, ensure_ascii=False))
    print(f"output={args.output_json}")


if __name__ == "__main__":
    main()
