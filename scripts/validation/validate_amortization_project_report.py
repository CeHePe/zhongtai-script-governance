from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill

from _project_root import find_project_root


ROOT = find_project_root(__file__)
TOLERANCE_YUAN = 0.01

REPORT_PREFIX = "1.1.4__摊销分析"
PLAN_FILE = "带资摊销、智能化整改、质效提升计划数.xlsx"
PROJECT_FILE = "项目查询.xlsx"
RATIO_FILE = "带资摊销比例配置.xlsx"
FINANCE_HISTORY_FILE = "财务云-实际发生数-导入结果0424.xls"
FINANCE_Q1_FILE = "摊销1-3月实际数-汇总.xlsx"

TYPE_FUNDED = "带资摊销"
TYPE_SMART = "智能化整改"
TYPE_QUALITY = "质效提升"

TYPE_MAP = {
    "带资进场": TYPE_FUNDED,
    "带资摊销": TYPE_FUNDED,
    "技术改造": TYPE_SMART,
    "智能化改造": TYPE_SMART,
    "智能化整改": TYPE_SMART,
    "美好家园": TYPE_QUALITY,
    "美好家园（投后）": TYPE_QUALITY,
    "存量精耕": TYPE_QUALITY,
    "标杆项目": TYPE_QUALITY,
    "质效提升": TYPE_QUALITY,
}


@dataclass(frozen=True)
class FieldSpec:
    position: int
    name: str


FIELDS = [
    FieldSpec(1, "序号"),
    FieldSpec(2, "区域"),
    FieldSpec(3, "项目"),
    FieldSpec(4, "项目状态"),
    FieldSpec(5, "条线"),
    FieldSpec(6, "我方持股比例"),
    FieldSpec(7, "分摊类型"),
    FieldSpec(8, "事项类型"),
    FieldSpec(9, "带资/整改金额（含税）"),
    FieldSpec(10, "带资/整改金额（不含税）"),
    FieldSpec(11, "剩余摊销金额（不含税）"),
    FieldSpec(12, "剩余发生金额（不含税）"),
    FieldSpec(13, "摊销年限"),
    FieldSpec(14, "摊销开始日期"),
    FieldSpec(15, "摊销结束日期"),
    FieldSpec(16, "往年累计-已摊销金额"),
    FieldSpec(17, "往年累计-已发生金额"),
    FieldSpec(18, "往年累计-还原金额"),
    FieldSpec(19, "当年累计-已摊销金额"),
    FieldSpec(20, "当年累计-已发生金额"),
    FieldSpec(21, "当年累计-还原金额"),
    FieldSpec(22, "一季度-已摊销金额"),
    FieldSpec(23, "一季度-已发生金额"),
    FieldSpec(24, "一季度-还原金额"),
    FieldSpec(25, "二季度-已摊销金额"),
    FieldSpec(26, "二季度-已发生金额"),
    FieldSpec(27, "二季度-还原金额"),
]


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Validate every field in the project amortization report.")
    parser.add_argument("--period", default="202605", help="Report period in YYYYMM format.")
    parser.add_argument(
        "--output",
        type=Path,
        help="Local-only validation workbook path.",
    )
    return parser.parse_args()


def normalize_period(value: object) -> str:
    text = str(value).strip().replace("/", "-")
    if len(text) >= 7 and text[4] == "-":
        return text[:7]
    if len(text) >= 6 and text[:6].isdigit():
        return f"{text[:4]}-{text[4:6]}"
    return text


def as_number(value: object) -> float:
    number = pd.to_numeric(pd.Series([value]), errors="coerce").iloc[0]
    return 0.0 if pd.isna(number) else float(number)


def equal_value(actual: object, expected: object, numeric: bool = False) -> tuple[bool, float | None]:
    if numeric:
        diff = as_number(actual) - as_number(expected)
        return abs(diff) < TOLERANCE_YUAN, diff
    left = "" if pd.isna(actual) else str(actual).strip()
    right = "" if pd.isna(expected) else str(expected).strip()
    return left == right, None


def load_report(period: str) -> tuple[Path, pd.DataFrame]:
    path = ROOT / f"{REPORT_PREFIX}{period}项目.xlsx"
    if not path.exists():
        raise FileNotFoundError(f"Report not found: {path.name}")
    report = pd.read_excel(path, header=[0, 1])
    if report.shape[1] != len(FIELDS):
        raise ValueError(f"Expected {len(FIELDS)} report fields, found {report.shape[1]}")
    return path, report


def load_plan(report_period: pd.Period) -> tuple[Path, pd.DataFrame, str]:
    path = ROOT / PLAN_FILE
    plan = pd.read_excel(path)
    months = sorted({normalize_period(value) for value in plan["数据年月"].dropna()})
    eligible = [month for month in months if pd.Period(month, freq="M") <= report_period]
    if not eligible:
        raise ValueError("No plan snapshot is on or before the report period")
    selected = eligible[-1]
    plan = plan[plan["数据年月"].map(normalize_period).eq(selected)].copy()
    return path, plan, selected


def load_ratio_map() -> dict[int, list[float]]:
    frame = pd.read_excel(ROOT / RATIO_FILE, header=1).dropna(subset=["分摊年限"])
    result: dict[int, list[float]] = {}
    for _, row in frame.iterrows():
        years = int(str(row["分摊年限"]).replace("年", ""))
        result[years] = [as_number(row.get(f"第{index}年_摊销比例")) for index in range(1, 7)]
    return result


def build_plan_mapping(report: pd.DataFrame, plan: pd.DataFrame) -> tuple[pd.DataFrame, pd.DataFrame]:
    keys = ["name", "type", "item", "amount", "years", "start", "end"]
    report_rows = pd.DataFrame(
        {
            "report_index": report.index,
            "excel_row": report.index + 3,
            "name": report.iloc[:, 2].astype(str).str.strip(),
            "type": report.iloc[:, 6].astype(str).str.strip(),
            "item": report.iloc[:, 7].astype(str).str.strip(),
            "amount": pd.to_numeric(report.iloc[:, 8], errors="coerce").round(6),
            "years": report.iloc[:, 12].astype(str).str.strip(),
            "start": report.iloc[:, 13].map(normalize_period),
            "end": report.iloc[:, 14].map(normalize_period),
        }
    )
    plan_rows = pd.DataFrame(
        {
            "plan_index": plan.index,
            "name": plan["项目名称"].astype(str).str.strip(),
            "type": plan["摊销类型"].astype(str).str.strip(),
            "item": plan["事项类型"].astype(str).str.strip(),
            "amount": pd.to_numeric(plan["计划投资金额（含税价）（单位：元）"], errors="coerce").round(6),
            "years": plan["分摊年限"].astype(str).str.strip(),
            "start": plan["摊销开始日期"].map(normalize_period),
            "end": plan["摊销结束日期"].map(normalize_period),
            "code": plan["立项编码"].astype(str).str.strip(),
            "status": plan["计划状态"].astype(str).str.strip(),
            "early_end": plan["提前结束年月"].map(normalize_period),
        }
    )
    report_rows["occurrence"] = report_rows.groupby(keys, dropna=False).cumcount()
    plan_rows["occurrence"] = plan_rows.groupby(keys, dropna=False).cumcount()
    mapped = report_rows.merge(plan_rows, on=keys + ["occurrence"], how="left", validate="one_to_one")
    return mapped, plan_rows


def plan_month_value(
    row: pd.Series,
    month: pd.Period,
    penetration: float,
    ratios: dict[int, list[float]],
    finance_actual_rows: pd.DataFrame,
) -> tuple[float, bool]:
    start = pd.Period(row["start"], freq="M")
    scheduled_end = pd.Period(row["end"], freq="M")
    status = str(row["status"])
    early_end = None
    if row.get("early_end") and row.get("early_end") != "nan":
        early_end = pd.Period(row["early_end"], freq="M")
    effective_end = min(scheduled_end, early_end) if early_end is not None else scheduled_end
    if month < start or month > effective_end:
        return 0.0, False
    terminal = month == effective_end
    amount = float(row["amount"])
    years = int(str(row["years"]).replace("年", ""))
    report_type = str(row["type"])
    if report_type == TYPE_SMART:
        return amount / years / 12 * penetration * 0.81, False
    if report_type == TYPE_QUALITY:
        return amount / years / 12 * penetration * 0.81 / 1.06, False

    def funded_raw(period: pd.Period) -> float:
        elapsed = (period.year - start.year) * 12 + period.month - start.month
        year_index = elapsed // 12
        ratio = ratios.get(years, [0.0] * 6)[year_index] if 0 <= year_index < 6 else 0.0
        return amount * ratio / 12

    if not terminal:
        return funded_raw(month) * penetration * 0.81 / 1.06, False

    prior_year_end = pd.Timestamp(f"{effective_end.year - 1}-12-31 23:59:59")
    actual_raw = finance_actual_rows.loc[
        finance_actual_rows["code"].eq(row["code"])
        & finance_actual_rows["type"].eq(TYPE_FUNDED)
        & finance_actual_rows["date"].ge(start.start_time)
        & finance_actual_rows["date"].le(prior_year_end),
        "raw",
    ].sum()
    prior_plan_raw = sum(funded_raw(period) for period in pd.period_range(start, effective_end - 1, freq="M"))
    value = actual_raw * 0.81 * penetration - prior_plan_raw * 0.81 / 1.06 * penetration
    return float(value), False


def calculate_plan_groups(
    plan_rows: pd.DataFrame,
    project: pd.DataFrame,
    report_period: pd.Period,
    ratios: dict[int, list[float]],
    finance_actual_rows: pd.DataFrame,
) -> pd.DataFrame:
    penetration = (
        project.drop_duplicates("立项编码").set_index("立项编码")["穿透比例"].to_dict()
    )
    periods = {
        "prior_plan": pd.period_range("2000-01", f"{report_period.year - 1}-12", freq="M"),
        "current_plan": pd.period_range(f"{report_period.year}-01", report_period, freq="M"),
        "q1_plan": pd.period_range(f"{report_period.year}-01", f"{report_period.year}-03", freq="M"),
        "q2_plan": pd.period_range(f"{report_period.year}-04", report_period, freq="M"),
    }
    records: list[dict[str, object]] = []
    for _, row in plan_rows.iterrows():
        record: dict[str, object] = {"code": row["code"], "type": row["type"]}
        row_penetration = as_number(penetration.get(row["code"], 1.0))
        for label, months in periods.items():
            total = 0.0
            blocked = False
            for month in months:
                value, terminal = plan_month_value(
                    row, month, row_penetration, ratios, finance_actual_rows
                )
                if terminal:
                    blocked = True
                elif not pd.isna(value):
                    total += float(value)
            record[label] = total
            record[f"{label}_blocked"] = blocked
        records.append(record)
    frame = pd.DataFrame(records)
    value_columns = list(periods)
    blocked_columns = [f"{column}_blocked" for column in value_columns]
    grouped_values = frame.groupby(["code", "type"], as_index=False)[value_columns].sum()
    grouped_blocked = frame.groupby(["code", "type"], as_index=False)[blocked_columns].max()
    return grouped_values.merge(grouped_blocked, on=["code", "type"], how="outer")


def load_finance_actual_rows() -> pd.DataFrame:
    parts: list[pd.DataFrame] = []

    history = pd.read_excel(ROOT / FINANCE_HISTORY_FILE)
    history["code"] = history["sub_project_code"].fillna(history["project_code"]).astype(str).str.strip()
    history["type"] = history["share_type_name"].map(TYPE_MAP)
    history["date"] = pd.to_datetime(history["approval_time"], errors="coerce")
    history["raw"] = pd.to_numeric(history["no_tax_dist_amount"], errors="coerce").fillna(0)
    parts.append(history[["code", "type", "date", "raw"]].assign(source="finance_history"))

    q1 = pd.read_excel(ROOT / FINANCE_Q1_FILE)
    q1 = q1[q1["是否分摊"].eq("是")].copy()
    q1["code"] = q1["立项编码"].astype(str).str.strip()
    q1["type"] = q1["摊销类型名称"].map(TYPE_MAP)
    q1["date"] = pd.to_datetime(q1["审批通过时间"], errors="coerce")
    q1["raw"] = pd.to_numeric(q1["分摊不含税金额"], errors="coerce").fillna(0)
    parts.append(q1[["code", "type", "date", "raw"]].assign(source="finance_q1"))

    actual = pd.concat(parts, ignore_index=True)
    actual = actual[actual["type"].notna() & actual["code"].ne("nan")].copy()
    return actual


def aggregate_finance_actuals(
    actual: pd.DataFrame, project: pd.DataFrame, report_period: pd.Period
) -> pd.DataFrame:
    penetration = project.drop_duplicates("立项编码").set_index("立项编码")["穿透比例"].to_dict()
    actual = actual.copy()
    actual["value"] = actual["raw"] * 0.81 * actual["code"].map(penetration).fillna(1.0)
    prior_end = pd.Timestamp(f"{report_period.year - 1}-12-31 23:59:59")
    q1_end = pd.Timestamp(f"{report_period.year}-03-31 23:59:59")
    q2_start = pd.Timestamp(f"{report_period.year}-04-01")
    report_end = report_period.end_time
    actual["prior_actual"] = actual["value"].where(actual["date"] <= prior_end, 0.0)
    actual["q1_actual"] = actual["value"].where(
        (actual["date"] > prior_end) & (actual["date"] <= q1_end), 0.0
    )
    actual["q2_actual"] = actual["value"].where(
        (actual["date"] >= q2_start) & (actual["date"] <= report_end), 0.0
    )
    actual["current_actual"] = actual["q1_actual"] + actual["q2_actual"]
    return actual.groupby(["code", "type"], as_index=False)[
        ["prior_actual", "current_actual", "q1_actual", "q2_actual"]
    ].sum()


def add_detail(
    details: list[dict[str, object]],
    excel_row: int,
    code: str,
    field: FieldSpec,
    actual: object,
    expected: object,
    status: str,
    reason: str,
    diff: float | None = None,
) -> None:
    details.append(
        {
            "Excel行": excel_row,
            "立项编码": code,
            "字段序号": field.position,
            "字段": field.name,
            "报表值": actual,
            "期望值": expected,
            "差异": diff,
            "状态": status,
            "说明": reason,
        }
    )


def validate(
    period: str,
) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame, dict[str, object]]:
    report_period = pd.Period(f"{period[:4]}-{period[4:6]}", freq="M")
    report_path, report = load_report(period)
    plan_path, plan, plan_snapshot = load_plan(report_period)
    project_path = ROOT / PROJECT_FILE
    project = pd.read_excel(project_path)
    project["立项编码"] = project["立项编码"].astype(str).str.strip()
    query = project.drop_duplicates("立项编码").set_index("立项编码")
    mapped, plan_rows = build_plan_mapping(report, plan)
    finance_actual_rows = load_finance_actual_rows()
    plan_groups = calculate_plan_groups(
        plan_rows, project, report_period, load_ratio_map(), finance_actual_rows
    )
    finance_groups = aggregate_finance_actuals(finance_actual_rows, project, report_period)
    mapped = mapped.merge(plan_groups, on=["code", "type"], how="left")
    mapped = mapped.merge(finance_groups, on=["code", "type"], how="left")
    for column in ["prior_actual", "current_actual", "q1_actual", "q2_actual"]:
        mapped[column] = mapped[column].fillna(0.0)

    details: list[dict[str, object]] = []
    period_fields = {
        16: ("prior_plan", "prior_plan_blocked"),
        19: ("current_plan", "current_plan_blocked"),
        22: ("q1_plan", "q1_plan_blocked"),
        25: ("q2_plan", "q2_plan_blocked"),
    }
    actual_fields = {17: "prior_actual", 20: "current_actual", 23: "q1_actual", 26: "q2_actual"}

    for _, row in mapped.iterrows():
        report_index = int(row["report_index"])
        excel_row = int(row["excel_row"])
        code = "" if pd.isna(row.get("code")) else str(row["code"])
        if not code:
            for field in FIELDS:
                add_detail(
                    details,
                    excel_row,
                    code,
                    field,
                    report.iloc[report_index, field.position - 1],
                    None,
                    "阻塞",
                    "无法匹配最近一期计划台账行",
                )
            continue

        project_row = query.loc[code]
        base_expected = {
            1: report_index + 1,
            2: project_row["所属区域"],
            3: project_row["项目名称"],
            4: project_row["项目状态"],
            5: project_row["业务属性"],
            6: project_row["穿透比例"],
            7: row["type"],
            8: row["item"],
            9: row["amount"],
            10: float(row["amount"]) / 1.06,
            13: row["years"],
            14: row["start"],
            15: row["end"],
        }
        numeric_base = {1, 6, 9, 10}
        for position, expected in base_expected.items():
            field = FIELDS[position - 1]
            actual = report.iloc[report_index, position - 1]
            if position in {14, 15}:
                actual = normalize_period(actual)
            passed, diff = equal_value(actual, expected, position in numeric_base)
            add_detail(
                details,
                excel_row,
                code,
                field,
                actual,
                expected,
                "通过" if passed else "失败",
                "最近季度计划台账/项目查询映射" if passed else "与最近季度计划台账/项目查询不一致",
                diff,
            )

        for position, (value_column, blocked_column) in period_fields.items():
            field = FIELDS[position - 1]
            actual = report.iloc[report_index, position - 1]
            expected = row[value_column]
            blocked = bool(row[blocked_column])
            if blocked:
                add_detail(
                    details,
                    excel_row,
                    code,
                    field,
                    actual,
                    expected,
                    "阻塞",
                    "该项目类型包含结束月/提前结束月，收口公式依赖完整财务云+业财累计发生数",
                )
            else:
                passed, diff = equal_value(actual, expected, True)
                add_detail(
                    details,
                    excel_row,
                    code,
                    field,
                    actual,
                    expected,
                    "通过" if passed else "失败",
                    "按指标清单逐月复算并按项目+分摊类型汇总",
                    diff,
                )

        for position, value_column in actual_fields.items():
            field = FIELDS[position - 1]
            actual = report.iloc[report_index, position - 1]
            expected = row[value_column]
            finance_pass, diff = equal_value(actual, expected, True)
            reason = (
                "财务云A为全量快照，缺少4至5月记录按0；"
                "业财B按用户确认视作0"
            )
            status = "通过" if finance_pass else "失败"
            add_detail(details, excel_row, code, field, actual, expected, status, reason, diff)

        internal_expected = {
            11: as_number(report.iloc[report_index, 9])
            - as_number(report.iloc[report_index, 15])
            - as_number(report.iloc[report_index, 18]),
            12: as_number(report.iloc[report_index, 9])
            - as_number(report.iloc[report_index, 16])
            - as_number(report.iloc[report_index, 19]),
            18: as_number(report.iloc[report_index, 16]) - as_number(report.iloc[report_index, 15]),
            21: as_number(report.iloc[report_index, 19]) - as_number(report.iloc[report_index, 18]),
            24: as_number(report.iloc[report_index, 22]) - as_number(report.iloc[report_index, 21]),
            27: as_number(report.iloc[report_index, 25]) - as_number(report.iloc[report_index, 24]),
        }
        for position, expected in internal_expected.items():
            field = FIELDS[position - 1]
            actual = report.iloc[report_index, position - 1]
            passed, diff = equal_value(actual, expected, True)
            add_detail(
                details,
                excel_row,
                code,
                field,
                actual,
                expected,
                "通过" if passed else "失败",
                "报表内部恒等式",
                diff,
            )

    detail = pd.DataFrame(details).sort_values(["字段序号", "Excel行"])
    summary = (
        detail.groupby(["字段序号", "字段", "状态"], as_index=False)
        .size()
        .pivot(index=["字段序号", "字段"], columns="状态", values="size")
        .fillna(0)
        .reset_index()
    )
    for column in ["通过", "失败", "阻塞"]:
        if column not in summary:
            summary[column] = 0
        summary[column] = summary[column].astype(int)
    summary["总行数"] = summary[["通过", "失败", "阻塞"]].sum(axis=1)
    summary["字段结论"] = np.select(
        [summary["失败"].gt(0), summary["阻塞"].gt(0)],
        ["失败", "阻塞"],
        default="通过",
    )
    summary = summary[["字段序号", "字段", "总行数", "通过", "失败", "阻塞", "字段结论"]]

    source_rows = [
        {"来源": "待测报表", "文件": report_path.name, "状态": "已读取", "说明": f"期间={period}, 维度=项目"},
        {"来源": "计划台账", "文件": plan_path.name, "状态": "已读取", "说明": f"按用户确认取最近季度快照={plan_snapshot}"},
        {"来源": "项目主数据", "文件": project_path.name, "状态": "已读取", "说明": "区域、项目状态、条线、穿透比例"},
        {"来源": "摊销比例", "文件": RATIO_FILE, "状态": "已读取", "说明": "带资计划分年比例"},
        {"来源": "财务云历史实际", "文件": FINANCE_HISTORY_FILE, "状态": "已读取", "说明": "财务云A路径"},
        {"来源": "财务云本年实际", "文件": FINANCE_Q1_FILE, "状态": "已读取/全量快照", "说明": "缺少4至5月记录按0"},
        {"来源": "业财实际", "文件": "", "状态": "按0处理", "说明": "用户确认业财B不存在"},
        {
            "来源": "二季度实际",
            "文件": FINANCE_Q1_FILE,
            "状态": "已覆盖/按0",
            "说明": f"财务云A为全量快照，{report_period.year}-04至{report_period}无记录即为0",
        },
    ]
    metadata = {
        "report_rows": len(report),
        "plan_snapshot": plan_snapshot,
        "unmatched_report_rows": int(mapped["plan_index"].isna().sum()),
        "failed_cells": int(detail["状态"].eq("失败").sum()),
        "blocked_cells": int(detail["状态"].eq("阻塞").sum()),
    }
    return summary, detail, pd.DataFrame(source_rows), metadata


def format_output(path: Path) -> None:
    wb = load_workbook(path)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    for ws in wb.worksheets:
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions
        ws.sheet_view.showGridLines = False
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = Font(color="FFFFFF", bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")
        for column in ws.columns:
            values = [str(cell.value or "") for cell in column[:200]]
            width = min(max(max(map(len, values), default=8) + 2, 10), 60)
            ws.column_dimensions[column[0].column_letter].width = width
    wb.save(path)


def main() -> None:
    args = parse_args()
    output = args.output or ROOT / "output" / "spreadsheet" / f"amortization_project_{args.period}_validation.xlsx"
    output.parent.mkdir(parents=True, exist_ok=True)
    summary, detail, sources, metadata = validate(args.period)
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary.to_excel(writer, sheet_name="字段汇总", index=False)
        detail.to_excel(writer, sheet_name="逐单元格明细", index=False)
        detail[detail["状态"].ne("通过")].to_excel(writer, sheet_name="失败及阻塞", index=False)
        sources.to_excel(writer, sheet_name="数据依赖", index=False)
    format_output(output)
    print(f"output={output}")
    print(f"report_rows={metadata['report_rows']}")
    print(f"plan_snapshot={metadata['plan_snapshot']}")
    print(f"unmatched_report_rows={metadata['unmatched_report_rows']}")
    print(f"failed_cells={metadata['failed_cells']}")
    print(f"blocked_cells={metadata['blocked_cells']}")
    print(summary.to_string(index=False))


if __name__ == "__main__":
    main()
