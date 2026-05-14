from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from _project_root import find_project_root
from typing import Dict, Iterable, List, Optional, Tuple

import math
import pandas as pd
from openpyxl import load_workbook


ROOT = find_project_root(__file__)
REPORT_YEAR = 2025
REPORT_MONTH = 12
LOGIC_TOL = 1e-6
FACT_TOL = 1e-6


@dataclass
class PlanRow:
    project_code: str
    project_name: str
    plan_status: str
    years: int
    amount_tax_yuan: float
    start_period: pd.Period
    end_period: pd.Period
    early_end_period: Optional[pd.Period]
    penetration: float


def find_file(contains_all: Iterable[str], suffix: str) -> Path:
    contains_all = list(contains_all)
    for path in ROOT.iterdir():
        if not path.is_file():
            continue
        if path.suffix.lower() != suffix.lower():
            continue
        name = path.name
        if all(token in name for token in contains_all):
            return path
    raise FileNotFoundError(f"未找到文件: {contains_all} {suffix}")


def parse_period(value: object) -> Optional[pd.Period]:
    if pd.isna(value):
        return None
    text = str(value).strip()
    if not text or text.lower() == "nan":
        return None
    text = text.replace("/", "-")
    if len(text) == 7:
        return pd.Period(text, freq="M")
    if len(text) == 10:
        return pd.Period(text[:7], freq="M")
    return pd.Period(text, freq="M")


def clean_number(value: object, default: float = 0.0) -> float:
    if pd.isna(value):
        return default
    return float(value)


def year_ratio_index(month: pd.Period, start: pd.Period) -> int:
    delta = (month.year - start.year) * 12 + (month.month - start.month)
    return delta // 12 + 1


def monthly_plan_raw_wanyuan(row: PlanRow, month: pd.Period, ratio_map: Dict[int, Dict[int, float]]) -> float:
    if month < row.start_period:
        return 0.0
    schedule_end = row.end_period
    if row.plan_status in {"提前结束", "已终止"} and row.early_end_period is not None:
        schedule_end = min(schedule_end, row.early_end_period)
    if month > schedule_end:
        return 0.0
    idx = year_ratio_index(month, row.start_period)
    ratio = ratio_map.get(row.years, {}).get(idx, 0.0)
    if ratio == 0:
        return 0.0
    return row.amount_tax_yuan / 10000.0 * ratio / 12.0


def prior_actual_raw_wanyuan(actual_df: pd.DataFrame, project_code: str, start_period: pd.Period) -> float:
    start_ts = start_period.to_timestamp()
    cutoff = pd.Timestamp(f"{REPORT_YEAR-1}-12-31")
    sub = actual_df[
        (actual_df["project_code_norm"] == project_code)
        & (actual_df["approval_time"] >= start_ts)
        & (actual_df["approval_time"] <= cutoff)
    ]
    return sub["no_tax_dist_amount"].fillna(0).sum() / 10000.0


def current_year_actual_raw_wanyuan(actual_df: pd.DataFrame, project_code: str) -> float:
    start_ts = pd.Timestamp(f"{REPORT_YEAR}-01-01")
    end_ts = pd.Timestamp(f"{REPORT_YEAR}-12-31")
    sub = actual_df[
        (actual_df["project_code_norm"] == project_code)
        & (actual_df["approval_time"] >= start_ts)
        & (actual_df["approval_time"] <= end_ts)
    ]
    return sub["no_tax_dist_amount"].fillna(0).sum() / 10000.0


def compute_row_month_metric(
    row: PlanRow,
    month: pd.Period,
    ratio_map: Dict[int, Dict[int, float]],
    actual_df: pd.DataFrame,
) -> Tuple[float, str]:
    status = row.plan_status
    penetration = row.penetration
    ordinary_raw = monthly_plan_raw_wanyuan(row, month, ratio_map)

    if status == "正常":
        if month < row.start_period or month > row.end_period:
            return 0.0, "不在摊销期间"
        if month < row.end_period:
            metric = ordinary_raw * 10000.0 * 0.81 / 1.06 * penetration
            return metric, "普通月"
        prior_plan_raw = 0.0
        jan_to_prev_raw = 0.0
        for m in pd.period_range(row.start_period, f"{REPORT_YEAR-1}-12", freq="M"):
            prior_plan_raw += monthly_plan_raw_wanyuan(row, m, ratio_map)
        for m in pd.period_range(f"{REPORT_YEAR}-01", month - 1, freq="M"):
            jan_to_prev_raw += monthly_plan_raw_wanyuan(row, m, ratio_map)
        prior_actual_raw = prior_actual_raw_wanyuan(actual_df, row.project_code, row.start_period)
        metric = (
            prior_actual_raw * 10000.0 * 0.81 * penetration
            - (prior_plan_raw + jan_to_prev_raw) * 10000.0 * 0.81 / 1.06 * penetration
        )
        return metric, "摊销结束月"

    if status in {"提前结束", "已终止"}:
        if row.early_end_period is None:
            return 0.0, "缺提前结束年月"
        if month < row.start_period or month > row.early_end_period:
            return 0.0, "不在提前结束期间"
        if month < row.early_end_period:
            # 按指标清单当前文字执行，未额外补 /12。
            idx = year_ratio_index(month, row.start_period)
            ratio = ratio_map.get(row.years, {}).get(idx, 0.0)
            metric = row.amount_tax_yuan * ratio * 0.81 / 1.06 * penetration
            return metric, "提前结束前月份"
        prior_plan_raw = 0.0
        jan_to_prev_raw = 0.0
        for m in pd.period_range(row.start_period, f"{REPORT_YEAR-1}-12", freq="M"):
            prior_plan_raw += monthly_plan_raw_wanyuan(row, m, ratio_map)
        for m in pd.period_range(f"{REPORT_YEAR}-01", month - 1, freq="M"):
            jan_to_prev_raw += monthly_plan_raw_wanyuan(row, m, ratio_map)
        prior_actual_raw = prior_actual_raw_wanyuan(actual_df, row.project_code, row.start_period)
        metric = (
            prior_actual_raw * 10000.0 * 0.81 * penetration
            - (prior_plan_raw + jan_to_prev_raw) * 10000.0 * 0.81 / 1.06 * penetration
        )
        return metric, "提前结束月"

    return 0.0, f"未知计划状态:{status}"


def load_indicator_logic() -> Dict[str, str]:
    path = find_file(["JKS_", "指标清单"], ".xlsx")
    wb = load_workbook(path, data_only=False)
    ws = wb[wb.sheetnames[0]]
    return {
        "row10_logic": str(ws.cell(10, 13).value or ""),
        "row31_logic": str(ws.cell(31, 13).value or ""),
        "row52_logic": str(ws.cell(52, 13).value or ""),
        "row73_logic": str(ws.cell(73, 13).value or ""),
    }


def load_ratio_map() -> Dict[int, Dict[int, float]]:
    path = find_file(["带资", "比例配置"], ".xlsx")
    df = pd.read_excel(path)
    ratio_map: Dict[int, Dict[int, float]] = {}
    for _, row in df.iloc[1:].iterrows():
        year_text = str(row.iloc[0]).strip()
        if not year_text or year_text.lower() == "nan":
            continue
        years = int(year_text.replace("年", ""))
        ratio_map[years] = {}
        for idx in range(1, 7):
            value = row.iloc[idx]
            if pd.notna(value):
                ratio_map[years][idx] = float(value)
    return ratio_map


def load_project_query() -> pd.DataFrame:
    path = find_file(["项目查询"], ".xlsx")
    df = pd.read_excel(path)
    df["立项编码"] = df["立项编码"].astype(str).str.strip()
    return df


def load_plan_rows(project_query: pd.DataFrame) -> List[PlanRow]:
    path = find_file(["带资", "智能化", "质效提升计划数"], ".xlsx")
    df = pd.read_excel(path)
    df = df[df["摊销类型"].eq("带资摊销")].copy()
    query_map = (
        project_query[["立项编码", "穿透比例"]]
        .drop_duplicates(subset=["立项编码"])
        .set_index("立项编码")["穿透比例"]
        .to_dict()
    )
    rows: List[PlanRow] = []
    for _, r in df.iterrows():
        code = str(r["立项编码"]).strip()
        penetration = clean_number(query_map.get(code, 1.0), 1.0)
        rows.append(
            PlanRow(
                project_code=code,
                project_name=str(r["项目名称"]).strip(),
                plan_status=str(r["计划状态"]).strip(),
                years=int(str(r["分摊年限"]).replace("年", "").strip()),
                amount_tax_yuan=clean_number(r["计划投资金额（含税价）（单位：元）"]),
                start_period=parse_period(r["摊销开始日期"]),
                end_period=parse_period(r["摊销结束日期"]),
                early_end_period=parse_period(r["提前结束年月"]),
                penetration=penetration,
            )
        )
    return rows


def load_report() -> pd.DataFrame:
    path = find_file(["管报归母净利润202512项目"], ".xlsx")
    df = pd.read_excel(path)
    df["项目"] = df["项目"].astype(str).str.strip()
    return df


def load_actuals() -> pd.DataFrame:
    path = find_file(["实际发生数", "0424"], ".xls")
    df = pd.read_excel(path, sheet_name=0)
    code = df["sub_project_code"].fillna(df["project_code"]).astype(str).str.strip()
    df["project_code_norm"] = code
    df["approval_time"] = pd.to_datetime(df["approval_time"])
    mask = (df["share_type_code"].astype(str) == "TX01") | (df["share_type_name"].astype(str) == "带资进场")
    return df[mask].copy()


def load_reconciliation() -> pd.DataFrame:
    path = find_file(["导入整理 V2"], ".xlsx")
    df = pd.read_excel(path, sheet_name=1, header=1)
    df = df.rename(
        columns={
            df.columns[18]: "project_code",
            df.columns[19]: "region_name",
            df.columns[20]: "project_name",
            df.columns[21]: "metric_type",
            df.columns[22]: "planned_raw_tax_wanyuan",
            df.columns[23]: "fact_plan_wanyuan",
            df.columns[24]: "fact_actual_wanyuan",
            df.columns[25]: "fact_restore_wanyuan",
        }
    )
    df["project_code"] = df["project_code"].astype(str).str.strip()
    df["metric_type"] = df["metric_type"].astype(str).str.strip()
    df = df[(df["metric_type"] == "带资") & (df["project_code"] != "nan")].copy()
    return df.groupby("project_code", as_index=False).agg(
        {
            "region_name": "first",
            "project_name": "first",
            "planned_raw_tax_wanyuan": "sum",
            "fact_plan_wanyuan": "sum",
            "fact_actual_wanyuan": "sum",
            "fact_restore_wanyuan": "sum",
        }
    )


def build_logic_results(
    plan_rows: List[PlanRow],
    report_df: pd.DataFrame,
    actual_df: pd.DataFrame,
    ratio_map: Dict[int, Dict[int, float]],
) -> pd.DataFrame:
    report_map = report_df.set_index("项目")
    months_2025 = list(pd.period_range("2025-01", "2025-12", freq="M"))
    by_project: Dict[str, Dict[str, object]] = {}

    for row in plan_rows:
        total = 0.0
        month_tags: List[str] = []
        for month in months_2025:
            value, tag = compute_row_month_metric(row, month, ratio_map, actual_df)
            total += value
            if abs(value) > LOGIC_TOL:
                month_tags.append(f"{month}:{tag}")
        item = by_project.setdefault(
            row.project_code,
            {
                "project_code": row.project_code,
                "project_name": row.project_name,
                "logic_expected_plan": 0.0,
                "logic_tags": [],
                "penetration": row.penetration,
                "plan_statuses": set(),
                "row_count": 0,
            },
        )
        item["logic_expected_plan"] += total
        item["logic_tags"].extend(month_tags)
        item["plan_statuses"].add(row.plan_status)
        item["row_count"] += 1

    rows = []
    report_plan_col = " 带资摊销调整计划数"
    report_actual_col = "带资摊销调整发生数"
    for code, item in by_project.items():
        report_plan = clean_number(report_map.at[code, report_plan_col]) if code in report_map.index else math.nan
        report_actual = clean_number(report_map.at[code, report_actual_col]) if code in report_map.index else math.nan
        diff = item["logic_expected_plan"] - report_plan if not math.isnan(report_plan) else math.nan
        current_year_actual = current_year_actual_raw_wanyuan(actual_df, code)
        rows.append(
            {
                "project_code": code,
                "project_name": item["project_name"],
                "report_plan": report_plan,
                "report_actual": report_actual,
                "logic_expected_plan": item["logic_expected_plan"],
                "logic_diff": diff,
                "logic_pass": (not math.isnan(diff)) and abs(diff) <= LOGIC_TOL,
                "logic_reason": "一致" if (not math.isnan(diff) and abs(diff) <= LOGIC_TOL) else "与最新指标清单逻辑不一致",
                "logic_trace": "；".join(item["logic_tags"]),
                "penetration": item["penetration"],
                "plan_statuses": "、".join(sorted(item["plan_statuses"])),
                "plan_row_count": item["row_count"],
                "current_year_actual_raw_wanyuan": current_year_actual,
            }
        )
    return pd.DataFrame(rows).sort_values(["logic_pass", "project_code"], ascending=[True, True])


def append_fact_results(logic_df: pd.DataFrame, recon_df: pd.DataFrame) -> pd.DataFrame:
    merged = logic_df.merge(recon_df, on="project_code", how="left", suffixes=("", "_recon"))
    merged["fact_expected_restore_metric"] = merged["fact_restore_wanyuan"].fillna(0) * 10000.0 * merged["penetration"]
    merged["report_restore_metric"] = merged["report_actual"].fillna(0) - merged["report_plan"].fillna(0)
    merged["fact_diff"] = merged["report_restore_metric"] - merged["fact_expected_restore_metric"]
    merged["fact_blocked"] = merged["fact_restore_wanyuan"].isna()
    merged["fact_pass"] = (~merged["fact_blocked"]) & (merged["fact_diff"].abs() <= FACT_TOL)
    merged["fact_reason"] = merged.apply(classify_fact_reason, axis=1)
    return merged


def classify_fact_reason(row: pd.Series) -> str:
    if bool(row["fact_blocked"]):
        return "摊销模板B-按项目统计缺少带资事实行"
    if bool(row["fact_pass"]):
        return "模板还原金额与报表发生-计划一致"
    if bool(row["logic_pass"]):
        return "技术实现已符合指标清单，但模板还原金额不能支持该年度结果；更像周期恒等式被当作年度恒等式使用"
    return "逻辑测试未通过，暂不据此下结论"


def build_summary(result_df: pd.DataFrame, indicator_logic: Dict[str, str]) -> pd.DataFrame:
    total = len(result_df)
    logic_fail = int((~result_df["logic_pass"]).sum())
    fact_blocked = int(result_df["fact_blocked"].sum())
    fact_fail = int((~result_df["fact_pass"] & ~result_df["fact_blocked"]).sum())
    fact_fail_after_logic_pass = int((result_df["logic_pass"] & ~result_df["fact_pass"] & ~result_df["fact_blocked"]).sum())
    rows = [
        {"项目": "复测项目数", "数值": total},
        {"项目": "逻辑测试通过数", "数值": total - logic_fail},
        {"项目": "逻辑测试不通过数", "数值": logic_fail},
        {"项目": "还原金额事实测试通过数", "数值": int(result_df["fact_pass"].sum())},
        {"项目": "还原金额事实测试不通过数", "数值": fact_fail},
        {"项目": "还原金额事实测试阻塞数", "数值": fact_blocked},
        {"项目": "逻辑已通过但还原金额不通过数", "数值": fact_fail_after_logic_pass},
        {"项目": "指标清单_序号9", "数值": indicator_logic["row10_logic"]},
        {"项目": "指标清单_序号30", "数值": indicator_logic["row31_logic"]},
        {"项目": "指标清单_序号51", "数值": indicator_logic["row52_logic"]},
        {"项目": "指标清单_序号72", "数值": indicator_logic["row73_logic"]},
    ]
    return pd.DataFrame(rows)


def build_client_report(result_df: pd.DataFrame) -> pd.DataFrame:
    issues = result_df[result_df["logic_pass"] & ~result_df["fact_pass"] & ~result_df["fact_blocked"]].copy()
    issues = issues.sort_values("fact_diff", key=lambda s: s.abs(), ascending=False)
    return issues[
        [
            "project_code",
            "project_name",
            "report_plan",
            "report_actual",
            "report_restore_metric",
            "fact_restore_wanyuan",
            "fact_expected_restore_metric",
            "fact_diff",
            "fact_reason",
        ]
    ]


def build_special_cases(result_df: pd.DataFrame) -> pd.DataFrame:
    sample_codes = ["D30202208290101R", "D3020230170101R", "P30202312003301R"]
    sub = result_df[result_df["project_code"].isin(sample_codes)].copy()
    return sub[
        [
            "project_code",
            "project_name",
            "report_plan",
            "logic_expected_plan",
            "logic_diff",
            "report_actual",
            "report_restore_metric",
            "fact_restore_wanyuan",
            "fact_expected_restore_metric",
            "fact_diff",
            "logic_pass",
            "fact_pass",
            "fact_reason",
        ]
    ]


def main() -> None:
    indicator_logic = load_indicator_logic()
    ratio_map = load_ratio_map()
    project_query = load_project_query()
    plan_rows = load_plan_rows(project_query)
    report_df = load_report()
    actual_df = load_actuals()
    recon_df = load_reconciliation()

    logic_df = build_logic_results(plan_rows, report_df, actual_df, ratio_map)
    result_df = append_fact_results(logic_df, recon_df)
    summary_df = build_summary(result_df, indicator_logic)
    client_df = build_client_report(result_df)
    sample_df = build_special_cases(result_df)

    output = ROOT / "带资项目双路径复测结果.xlsx"
    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="摘要", index=False)
        result_df.to_excel(writer, sheet_name="全量明细", index=False)
        client_df.to_excel(writer, sheet_name="甲方缺陷候选", index=False)
        sample_df.to_excel(writer, sheet_name="历史3项目", index=False)

    print(f"输出文件: {output}")
    print(summary_df.to_string(index=False))


if __name__ == "__main__":
    main()
