from __future__ import annotations

import argparse
import json
import re
from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string


TOLERANCE_YUAN = 0.01


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(
        description="Validate funded-input fields in the external-quality project report."
    )
    parser.add_argument("--report", type=Path, required=True)
    parser.add_argument("--projection-ledger", type=Path, required=True)
    parser.add_argument("--report-period", required=True)
    parser.add_argument("--report-column", default="AB")
    parser.add_argument("--output-json", type=Path, required=True)
    return parser.parse_args()


def normalize_text(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalize_period(value: Any) -> str:
    digits = re.sub(r"\D", "", normalize_text(value))
    return digits[:6] if len(digits) >= 6 else ""


def number(value: Any) -> float:
    if value in (None, ""):
        return 0.0
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def iter_values(path: Path, min_row: int):
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    sheet.reset_dimensions()
    yield from enumerate(sheet.iter_rows(min_row=min_row, values_only=True), min_row)


def validate(args: argparse.Namespace) -> dict[str, Any]:
    if not re.fullmatch(r"\d{6}", args.report_period):
        raise ValueError("--report-period must use YYYYMM")
    for path in (args.report, args.projection_ledger):
        if not path.exists():
            raise FileNotFoundError(path)

    report_col = column_index_from_string(args.report_column.upper()) - 1
    ledger_by_code: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for excel_row, row in iter_values(args.projection_ledger, 4):
        source_period = normalize_period(row[0] if len(row) > 0 else None)
        if source_period != args.report_period:
            continue
        code = normalize_text(row[2] if len(row) > 2 else None)
        if not code:
            continue
        ledger_by_code[code].append(
            {
                "excel_row": excel_row,
                "project_name": normalize_text(row[1] if len(row) > 1 else None),
                "value": number(row[16] if len(row) > 16 else None),
            }
        )

    details: list[dict[str, Any]] = []
    report_codes: set[str] = set()
    for excel_row, row in iter_values(args.report, 4):
        code = normalize_text(row[1] if len(row) > 1 else None)
        project_name = normalize_text(row[2] if len(row) > 2 else None)
        report_codes.add(code)
        actual = number(row[report_col] if len(row) > report_col else None)
        source_rows = ledger_by_code.get(code, [])
        expected = source_rows[0]["value"] if len(source_rows) == 1 else 0.0
        diff = actual - expected
        if len(source_rows) > 1:
            status = "阻塞"
            difference_type = "底表项目重复"
        elif abs(diff) < TOLERANCE_YUAN:
            status = "通过"
            difference_type = ""
        else:
            status = "失败"
            difference_type = "取值差异"
        ratio = None
        if abs(expected) >= TOLERANCE_YUAN:
            ratio = actual / expected
        details.append(
            {
                "报表行": excel_row,
                "项目编码": code,
                "项目名称": project_name,
                "字段列": args.report_column.upper(),
                "字段": "投模带资投入费用",
                "报表值": actual,
                "底表值": expected,
                "差异": diff,
                "报表值/底表值": ratio,
                "底表匹配行数": len(source_rows),
                "底表行": "、".join(str(item["excel_row"]) for item in source_rows),
                "差异类型": difference_type,
                "状态": status,
                "说明": (
                    "按项目编码和报表期间唯一匹配；底表缺行按0"
                    if len(source_rows) <= 1
                    else "同一项目编码和期间存在多行，未强制汇总"
                ),
            }
        )

    status_counts = Counter(item["状态"] for item in details)
    failure_ratios = [
        item["报表值/底表值"]
        for item in details
        if item["状态"] == "失败" and item["报表值/底表值"] is not None
    ]
    all_failures_are_double = bool(failure_ratios) and all(
        abs(value - 2.0) < 1e-9 for value in failure_ratios
    )
    source_only = []
    for code in sorted(set(ledger_by_code) - report_codes):
        for item in ledger_by_code[code]:
            source_only.append(
                {
                    "项目编码": code,
                    "项目名称": item["project_name"],
                    "底表值": item["value"],
                    "底表行": item["excel_row"],
                    "说明": "底表存在，但全量报表中没有该项目编码",
                }
            )

    return {
        "metadata": {
            "report_period": args.report_period,
            "report_column": args.report_column.upper(),
            "tolerance_yuan": TOLERANCE_YUAN,
            "missing_row_policy": "底表缺行按0",
            "match_rule": "项目编码+报表期间唯一匹配",
        },
        "summary": {
            "report_rows": len(details),
            "ledger_rows": sum(len(rows) for rows in ledger_by_code.values()),
            "ledger_codes": len(ledger_by_code),
            "ledger_duplicate_codes": sum(
                1 for rows in ledger_by_code.values() if len(rows) > 1
            ),
            "ledger_codes_not_in_report": len(set(ledger_by_code) - report_codes),
            "report_nonzero_rows": sum(
                1 for item in details if abs(item["报表值"]) >= TOLERANCE_YUAN
            ),
            "ledger_nonzero_rows": sum(
                1 for item in details if abs(item["底表值"]) >= TOLERANCE_YUAN
            ),
            "status_counts": dict(status_counts),
            "all_failures_are_exactly_double": all_failures_are_double,
        },
        "details": details,
        "source_only": source_only,
    }


def main() -> None:
    args = parse_args()
    result = validate(args)
    args.output_json.parent.mkdir(parents=True, exist_ok=True)
    args.output_json.write_text(
        json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8"
    )
    print(json.dumps(result["summary"], ensure_ascii=False))
    print(f"output={args.output_json}")


if __name__ == "__main__":
    main()
