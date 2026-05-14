from __future__ import annotations

from copy import copy
from pathlib import Path
from _project_root import find_project_root

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = find_project_root(__file__)


def u(text: str) -> str:
    return text.encode("utf-8").decode("unicode_escape")


REPORT_NAME = u(r"\u4e0a\u4e00\u5e74\u56de\u6b3e\u7387202512\u9879\u76ee.xlsx")
AGING_2024_NAME = u(r"\u4e1a\u52a1\u5e10\u9f84-\u5e74\u5ea6\u5206\u5e03202412.xlsx")
AGING_2025_NAME = u(r"\u4e1a\u52a1\u5e10\u9f84-\u5e74\u5ea6\u5206\u5e03202512.xlsx")

SOURCE_METRIC_HEADER = u(r"\u4e0a\u4e00\u5e74\u5e94\u6536\u5728\u5f53\u671f\u56de\u6b3e")
NEW_HEADERS = [
    u(r"2024-12\u7684 2024\u5e74"),
    u(r"2025-12\u7684 2024\u5e74"),
    u(r"\u4e0a\u4e00\u5e74\u5e94\u6536\u5728\u5f53\u671f\u56de\u6b3e_\u8ba1\u7b97\u503c"),
    u(r"\u4e0a\u4e00\u5e74\u5e94\u6536\u5728\u5f53\u671f\u56de\u6b3e_\u5dee\u5f02\u503c"),
]


def copy_cell_style(source, target) -> None:
    target.font = copy(source.font)
    target.fill = copy(source.fill)
    target.border = copy(source.border)
    target.alignment = copy(source.alignment)
    target.number_format = source.number_format
    target.protection = copy(source.protection)


def find_report_paths() -> tuple[Path, Path, Path]:
    report = ROOT / REPORT_NAME
    aging_2024 = ROOT / AGING_2024_NAME
    aging_2025 = ROOT / AGING_2025_NAME
    return report, aging_2024, aging_2025


def find_metric_column(ws) -> int:
    for col in range(1, ws.max_column + 1):
        if ws.cell(2, col).value == SOURCE_METRIC_HEADER:
            return col
    raise RuntimeError("Could not find source metric column in row 2.")


def find_or_insert_columns(ws, metric_col: int) -> tuple[int, int, int, int]:
    existing = [ws.cell(2, metric_col + offset).value for offset in range(1, 5)]
    if existing == NEW_HEADERS:
        return metric_col + 1, metric_col + 2, metric_col + 3, metric_col + 4

    ws.insert_cols(metric_col + 1, amount=4)
    new_cols = [metric_col + 1, metric_col + 2, metric_col + 3, metric_col + 4]

    for col in new_cols:
        ws.cell(1, col).value = None

    style_row_1_source = metric_col + 5
    style_row_2_source = metric_col
    style_row_3_source = metric_col

    for idx, col in enumerate(new_cols):
        ws.cell(2, col).value = NEW_HEADERS[idx]
        copy_cell_style(ws.cell(1, style_row_1_source), ws.cell(1, col))
        copy_cell_style(ws.cell(2, style_row_2_source), ws.cell(2, col))
        copy_cell_style(ws.cell(3, style_row_3_source), ws.cell(3, col))
        ws.column_dimensions[get_column_letter(col)].width = ws.column_dimensions[
            get_column_letter(style_row_3_source)
        ].width

    return tuple(new_cols)


def populate_formulas(ws, metric_col: int, col_2024: int, col_2025: int, calc_col: int, diff_col: int) -> None:
    report_metric_col_letter = get_column_letter(metric_col)
    col_2024_letter = get_column_letter(col_2024)
    col_2025_letter = get_column_letter(col_2025)
    calc_col_letter = get_column_letter(calc_col)

    source_2024 = f"'[{AGING_2024_NAME}]sheet1'!$B:$B"
    source_2024_val = f"'[{AGING_2024_NAME}]sheet1'!$M:$M"
    source_2025 = f"'[{AGING_2025_NAME}]sheet1'!$B:$B"
    source_2025_val = f"'[{AGING_2025_NAME}]sheet1'!$N:$N"

    for row in range(3, ws.max_row + 1):
        project_code = ws.cell(row, 4).value
        if project_code in (None, ""):
            continue

        ws.cell(row, col_2024).value = (
            f'=IFERROR(XLOOKUP($D{row},{source_2024},{source_2024_val},0),0)'
        )
        ws.cell(row, col_2025).value = (
            f'=IFERROR(XLOOKUP($D{row},{source_2025},{source_2025_val},0),0)'
        )
        ws.cell(row, calc_col).value = f"={col_2024_letter}{row}-{col_2025_letter}{row}"
        ws.cell(row, diff_col).value = f"={report_metric_col_letter}{row}-{calc_col_letter}{row}"

        copy_cell_style(ws.cell(row, metric_col), ws.cell(row, col_2024))
        copy_cell_style(ws.cell(row, metric_col), ws.cell(row, col_2025))
        copy_cell_style(ws.cell(row, metric_col), ws.cell(row, calc_col))
        copy_cell_style(ws.cell(row, metric_col), ws.cell(row, diff_col))


def main() -> None:
    report_path, _, _ = find_report_paths()
    wb = load_workbook(report_path)
    ws = wb[wb.sheetnames[0]]

    metric_col = find_metric_column(ws)
    col_2024, col_2025, calc_col, diff_col = find_or_insert_columns(ws, metric_col)
    populate_formulas(ws, metric_col, col_2024, col_2025, calc_col, diff_col)

    wb.save(report_path)
    print(report_path.name.encode("unicode_escape").decode())
    print(f"metric_col={metric_col}, inserted_cols={[col_2024, col_2025, calc_col, diff_col]}")


if __name__ == "__main__":
    main()
