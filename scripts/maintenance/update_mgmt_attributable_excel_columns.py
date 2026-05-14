from __future__ import annotations

from pathlib import Path
from _project_root import find_project_root

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


ROOT = find_project_root(__file__)


def u(text: str) -> str:
    return text.encode("utf-8").decode("unicode_escape")


ATTRIBUTABLE = u(r"\u7ba1\u62a5\u5f52\u6bcd\u51c0\u5229\u6da6")
PROJECT = u(r"\u9879\u76ee")
SPACE = u(r"\u7a7a\u95f4")
CALC_HEADER = u(r"\u8ba1\u7b97\u7ed3\u679c")
DIFF_HEADER = u(r"\u5dee\u5f02")


def find_workbook(*tokens: str) -> Path:
    matches = [
        path
        for path in ROOT.iterdir()
        if path.suffix.lower() == ".xlsx" and all(token in path.name for token in tokens)
    ]
    if len(matches) != 1:
        raise RuntimeError(f"Expected one workbook for {tokens}, got {len(matches)}")
    return matches[0]


def ensure_formula_columns(
    path: Path,
    attributable_header: str,
    calc_formula_builder,
) -> tuple[str, str]:
    wb = load_workbook(path)
    ws = wb.active

    headers = [ws.cell(1, col).value for col in range(1, ws.max_column + 1)]
    try:
        attributable_col = headers.index(attributable_header) + 1
    except ValueError as exc:
        raise RuntimeError(f"{path.name} missing header {attributable_header}") from exc

    calc_col = attributable_col + 1
    diff_col = attributable_col + 2
    calc_letter = get_column_letter(calc_col)
    diff_letter = get_column_letter(diff_col)
    attributable_letter = get_column_letter(attributable_col)

    ws.cell(1, calc_col).value = CALC_HEADER
    ws.cell(1, diff_col).value = DIFF_HEADER

    for row in range(2, ws.max_row + 1):
        ws.cell(row, calc_col).value = calc_formula_builder(row)
        ws.cell(row, diff_col).value = f"={calc_letter}{row}-{attributable_letter}{row}"

    wb.save(path)
    return calc_letter, diff_letter


def project_formula(row: int) -> str:
    return f"=F{row}-I{row}-J{row}-L{row}-N{row}+K{row}+M{row}+O{row}-P{row}+Q{row}-R{row}+S{row}"


def space_formula(row: int) -> str:
    return f"=B{row}-C{row}-E{row}-F{row}-H{row}-J{row}+G{row}+I{row}+K{row}-L{row}+M{row}-N{row}+O{row}+P{row}+Q{row}+R{row}-S{row}"


def main() -> None:
    project_path = find_workbook(ATTRIBUTABLE, PROJECT)
    space_path = find_workbook(ATTRIBUTABLE, SPACE)

    project_cols = ensure_formula_columns(project_path, ATTRIBUTABLE, project_formula)
    space_cols = ensure_formula_columns(space_path, ATTRIBUTABLE, space_formula)

    print(f"{project_path.name}: {project_cols[0]}={CALC_HEADER}, {project_cols[1]}={DIFF_HEADER}")
    print(f"{space_path.name}: {space_cols[0]}={CALC_HEADER}, {space_cols[1]}={DIFF_HEADER}")


if __name__ == "__main__":
    main()
