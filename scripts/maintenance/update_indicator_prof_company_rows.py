from __future__ import annotations

from copy import copy
from datetime import datetime
from pathlib import Path
from _project_root import find_project_root
import shutil

from openpyxl import load_workbook
from openpyxl.worksheet.cell_range import CellRange


ROOT = find_project_root(__file__)


def u(text: str) -> str:
    return text.encode("ascii").decode("unicode_escape")


METRIC_CURRENT = u(r"\u5355\u4e00\u4e1a\u6743\u5f53\u5e74\u5e94\u6536\u5728\u5f53\u671f\u672a\u5230\u8d26\u671f\u91d1\u989d")
METRIC_BEGINNING = u(r"\u5355\u4e00\u4e1a\u6743\u4e0a\u4e00\u5e74\u672b\u5e94\u6536\u672a\u5230\u8d26\u671f\u91d1\u989d")

ORG_SPACE = u(r"\u7a7a\u95f4")
PERIOD_CUMULATIVE = u(r"\u7d2f\u8ba1")
PERIOD_MONTHLY = u(r"\u5355\u6708")
MODE_MANUAL_LEDGER = u(r"\u624b\u5de5\u53f0\u8d26")
MODE_SYSTEM_CALC = u(r"\u7cfb\u7edf\u8ba1\u7b97")
PROF_LEDGER = u(r"\u4e13\u4e1a\u516c\u53f8\u7684\u56de\u6b3e\u8425\u6536\u6bd4\u8c03\u6574\u9879\u53f0\u8d26")
INDEX_LIBRARY = u(r"\u6307\u6807\u5e93")
ORG_TYPES = [
    u(r"\u5b89\u4fdd"),
    u(r"\u91d1\u9890"),
    u(r"\u91d1\u4ee4\u91d1\u5320"),
    u(r"\u97f5\u6db5"),
]
WRONG_ORG = u(r"\u5168\u989d")
CORRECT_ORG = u(r"\u91d1\u9890")
SPACE_TYPES = u(r"\u533a\u57df\u3001\u5b89\u4fdd\u3001\u91d1\u9890\u3001\u91d1\u4ee4\u91d1\u5320\u3001\u97f5\u6db5")


def copy_row_style(ws, src_row: int, dst_row: int) -> None:
    ws.row_dimensions[dst_row].height = ws.row_dimensions[src_row].height
    for col in range(1, ws.max_column + 1):
        src = ws.cell(src_row, col)
        dst = ws.cell(dst_row, col)
        if src.has_style:
            dst._style = copy(src._style)
        if src.number_format:
            dst.number_format = src.number_format
        if src.font:
            dst.font = copy(src.font)
        if src.fill:
            dst.fill = copy(src.fill)
        if src.border:
            dst.border = copy(src.border)
        if src.alignment:
            dst.alignment = copy(src.alignment)
        if src.protection:
            dst.protection = copy(src.protection)


def shifted_range(cell_range: CellRange, idx: int, amount: int) -> CellRange:
    new_range = copy(cell_range)
    if cell_range.max_row < idx:
        return new_range
    if cell_range.min_row >= idx:
        new_range.shift(row_shift=amount)
        return new_range
    if cell_range.min_row < idx <= cell_range.max_row:
        new_range.max_row += amount
        return new_range
    return new_range


def insert_rows_preserving_merges(ws, idx: int, amount: int, template_row: int) -> None:
    old_ranges = [copy(cell_range) for cell_range in ws.merged_cells.ranges]
    for cell_range in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(cell_range))

    ws.insert_rows(idx, amount)
    for row in range(idx, idx + amount):
        copy_row_style(ws, template_row, row)

    for cell_range in old_ranges:
        ws.merge_cells(str(shifted_range(cell_range, idx, amount)))


def has_rows(ws, metric: str, period: str) -> bool:
    found = {
        ws.cell(row, 5).value
        for row in range(2, ws.max_row + 1)
        if ws.cell(row, 4).value == metric and ws.cell(row, 6).value == period
    }
    return all(org in found for org in ORG_TYPES)


def find_space_row(ws, metric: str, period: str, start_row: int) -> int:
    for row in range(start_row, ws.max_row + 1):
        if (
            ws.cell(row, 4).value == metric
            and ws.cell(row, 5).value == ORG_SPACE
            and ws.cell(row, 6).value == period
        ):
            return row
    raise ValueError(f"Cannot find space row for {metric!r} {period!r}")


def find_metric_start(ws, metric: str) -> int:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 3).value == metric and ws.cell(row, 4).value == metric:
            return row
    raise ValueError(f"Cannot find metric start for {metric!r}")


def cumulative_logic(org: str, field_name: str) -> str:
    return (
        f"根据数据年月=报表年月+类型={org}，查询【{PROF_LEDGER}】，"
        f"取【{field_name}】字段值"
    )


def monthly_logic(org: str, metric: str) -> str:
    return (
        f"A=根据类型={org}+数据年月=报表年月+累计/单月=累计，查询【{INDEX_LIBRARY}】，"
        f"如果查询到，取【{metric}】字段值\n"
        f"B=根据类型={org}+数据年月=报表年月-1+累计/单月=累计，查询【{INDEX_LIBRARY}】，"
        f"如果查询到，取【{metric}】字段值\n"
        f"{metric}=A-B"
    )


def space_cumulative_logic(metric: str) -> str:
    return (
        f"根据类型={SPACE_TYPES}+期间+累计/单月=累计，查询【{INDEX_LIBRARY}】，"
        f"取【{metric}】字段值，并进行汇总\n"
        "如：出2025年8月份报表，所属期间选择2025年8月"
    )


def fill_prof_rows(ws, start_row: int, metric: str, period: str, field_name: str | None = None) -> None:
    for offset, org in enumerate(ORG_TYPES):
        row = start_row + offset
        ws.cell(row, 4).value = metric
        ws.cell(row, 5).value = org
        ws.cell(row, 6).value = period
        ws.cell(row, 7).value = None
        ws.cell(row, 8).value = None
        ws.cell(row, 10).value = None
        ws.cell(row, 12).value = None
        ws.cell(row, 14).value = None
        ws.cell(row, 15).value = None
        if period == PERIOD_CUMULATIVE:
            ws.cell(row, 9).value = MODE_MANUAL_LEDGER
            ws.cell(row, 11).value = PROF_LEDGER
            if field_name is None:
                raise ValueError("field_name is required for cumulative rows")
            ws.cell(row, 13).value = cumulative_logic(org, field_name)
        else:
            ws.cell(row, 9).value = MODE_SYSTEM_CALC
            ws.cell(row, 11).value = None
            ws.cell(row, 13).value = monthly_logic(org, metric)


def normalize_misread_org(ws) -> None:
    for row in range(2, ws.max_row + 1):
        if ws.cell(row, 4).value not in {METRIC_CURRENT, METRIC_BEGINNING}:
            continue
        if ws.cell(row, 5).value == WRONG_ORG:
            ws.cell(row, 5).value = CORRECT_ORG
        logic = ws.cell(row, 13).value
        if isinstance(logic, str):
            logic = logic.replace(f"类型={WRONG_ORG}", f"类型={CORRECT_ORG}")
            logic = logic.replace(
                u(r"\u533a\u57df\u3001\u5b89\u4fdd\u3001\u5168\u989d\u3001\u91d1\u4ee4\u91d1\u5320\u3001\u97f5\u6db5"),
                SPACE_TYPES,
            )
            ws.cell(row, 13).value = logic


def edit_metric(ws, metric: str, cumulative_field: str) -> None:
    start_row = find_metric_start(ws, metric)

    if not has_rows(ws, metric, PERIOD_CUMULATIVE):
        cumulative_space_row = find_space_row(ws, metric, PERIOD_CUMULATIVE, start_row)
        insert_rows_preserving_merges(ws, cumulative_space_row, len(ORG_TYPES), cumulative_space_row - 1)
        fill_prof_rows(ws, cumulative_space_row, metric, PERIOD_CUMULATIVE, cumulative_field)

    start_row = find_metric_start(ws, metric)
    cumulative_space_row = find_space_row(ws, metric, PERIOD_CUMULATIVE, start_row)
    ws.cell(cumulative_space_row, 13).value = space_cumulative_logic(metric)

    if not has_rows(ws, metric, PERIOD_MONTHLY):
        monthly_space_row = find_space_row(ws, metric, PERIOD_MONTHLY, start_row)
        insert_rows_preserving_merges(ws, monthly_space_row, len(ORG_TYPES), monthly_space_row - 1)
        fill_prof_rows(ws, monthly_space_row, metric, PERIOD_MONTHLY)


def resequence(ws) -> None:
    for row in range(2, ws.max_row + 1):
        is_inserted_prof_row = (
            ws.cell(row, 4).value in {METRIC_CURRENT, METRIC_BEGINNING}
            and ws.cell(row, 5).value in set(ORG_TYPES)
        )
        if ws.cell(row, 1).value is not None or is_inserted_prof_row:
            ws.cell(row, 1).value = row - 1


def main() -> None:
    path = next(path for path in ROOT.glob("*.xlsx") if path.name.startswith("JKS_"))
    backup = path.with_suffix(path.suffix + f".bak-{datetime.now():%Y%m%d%H%M%S}")
    shutil.copy2(path, backup)

    wb = load_workbook(path)
    ws = wb.active

    normalize_misread_org(ws)

    edit_metric(
        ws,
        METRIC_CURRENT,
        u(r"\u672a\u8fbe\u8d26\u671f\u91d1\u989d\uff1a\u672c\u5e74\u622a\u81f3\u5230\u5f53\u524d\u6708\u4efd").replace("e", ""),
    )
    edit_metric(
        ws,
        METRIC_BEGINNING,
        u(r"\u672a\u8fbe\u8d26\u671f\u91d1\u989d\uff1a\u672c\u5e74\u671f\u521d\u672a\u5230\u8d26\u671f\u91d1\u989d").replace("e", ""),
    )

    resequence(ws)
    ws.auto_filter.ref = f"A1:O{ws.max_row}"
    wb.save(path)
    print(f"updated={path}")
    print(f"backup={backup}")


if __name__ == "__main__":
    main()
